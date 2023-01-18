# 引用必要套件
import firebase_admin
from firebase_admin import credentials


#from firebase_admin import firestore
from firebase_admin import db
#import matplotlib.pyplot as plt
#import pandas as pd
#import numpy as np
import json
import datetime
from datetime import date
import tkinter as tk

# 引入訊息視窗模組
import tkinter.messagebox

# 引用私密金鑰
# path/to/serviceAccount.json 請用自己存放的路徑
cred = credentials.Certificate('/Users/home/Desktop/python/outdata-6457c-firebase-adminsdk-2miud-7b496d14a7 (1).json')

# 初始化firebase，注意不能重複初始化
firebase_admin.initialize_app(cred,{
	'databaseURL':'https://outdata-6457c.firebaseio.com'})



    



#0019
def geteveryday(begin_date,end_date):#計算日期天數###
    date_list = []
    begin_date = datetime.datetime.strptime(begin_date, "%Y%m%d")
    end_date = datetime.datetime.strptime(end_date,"%Y%m%d")
    while begin_date <= end_date:
        date_str = begin_date.strftime("%Y%m%d")
        date_list.append(date_str)
        begin_date += datetime.timedelta(days=1)
    return date_list
#0018
def isValidDate(datestr):#是否日期格式###
    try:
        date.fromisoformat(datestr)
    except:
        return False
    else:
        return True


#0016
def form_use(name):#excel 檔案有沒關###
    try:
        import openpyxl
            
        wb = openpyxl.load_workbook(name, data_only=True)
        wb.save(name)
    except:
        return False
    else:
        return True



def show_allform_window():#日期,序號,廠商名,資料
    
    window = tk.Tk()

    window.title('輸入表單')#程式上方的文字
    window.geometry("1300x700+25+25")
    window.resizable(False, False)#定義可不可以被使用者放大縮小視窗
    
    
    def form_search_only(temp,flag):#025 
        
        middle_clean()
        account_temp=[]
        path_form='account/'
        account_temp=(readdate(path_form))#表單資料
        path_form_name='inguire/'
        account_name_temp=(readdate(path_form_name))#表單資料+name

        title=[]
        for  hh in account_name_temp:
            title.append(hh) #表單標籤資料有名字 
        day_choose=geteveryday(temp[0],temp[1])#取得所有日子
        if flag==9:
            name=temp[2]
        second_temp=[]
        for i in range(len(day_choose)):
            for j in range(len(title)):
                if(day_choose[i] in title[j]):
                    second_temp.append(title[j])#日期篩選
        if flag==12 or flag==13:
            #print(second_temp)#p034b p035b
            #test_show_1("12")
            if len(second_temp) < 21:
                dict_form_list={}
                for i in range(len(second_temp)):
                    dict_form_list[i]=second_temp[i] #
                #03-03-04  #03-04-04
                middle_show_b(1,"表單列表",dict_form_list,flag)#34b(12 改flag)
            
            else:
                test_show_1("縮小選擇範圍")
        
        if flag==9:
            third_temp=[]#日期+名子篩選
            for i in range(len(second_temp)):
                if name in second_temp[i]:
                    third_temp.append(second_temp[i])
            test_show_1(len(third_temp))
            #p026
            trans_temp=[]
            last_toatl=0
            #savey=0
            temp=[]
            for i in range(len(third_temp)): #
                one_list= json.loads(account_temp.get(third_temp[i][0:11]))
                one3_list=one_list[2]
                dete_a=one_list[0] #日期
                if i ==0:
                    name=one_list[1]#客戶名
                for i in range(len(one3_list)) :  
                    if i==0:
                        temp.append(dete_a)  
                    else :
                        temp.append("")
                    
                    temp.append(one3_list[i][0])
                    temp.append(one3_list[i][1])
                    temp.append(one3_list[i][2])
                    temp.append(one3_list[i][3])
                    temp.append(one3_list[i][4])
                    last_toatl=last_toatl+int(one3_list[i][4])
                    trans_temp.append(temp)
                    temp=[]
            
            #print(last_toatl)   
            temp=["","---------------------","-----","-----","-----","-----"]
            trans_temp.append(temp)
            temp=["","總計","","","",last_toatl]
            trans_temp.append(temp)    
            right_show_b(trans_temp,name,last_toatl)#03-01-05
            form_print(trans_temp,name,9)#03-01-05a
            #
            #trans_temp=[]
            for i in range(len(form_search)):
                del form_search[0]
            
    def form_print(temp,name,flag): #p26 輸出excel
        #03-01-05a
        
        import os
        os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

        import openpyxl
        wb = openpyxl.load_workbook('form_use.xlsx', data_only=True)

        s1 = wb['sheet1']            # 開啟工作表 1
        s1['A1'].value = '客戶名稱'     # 儲存格 A1 內容為
        s1['B1'].value = name    # 儲存格 A2 內容為 
        #s1['C1'].value = '大賣'    # 儲存格 A3 內容為 
        
        wire_temp=[]# 電線名,成本,大賣,標籤
        temp_use=[]
        s1.cell(2,1).value ='日　期  '
        s1.cell(2,2).value ='產　品　名　稱 '
        s1.cell(2,3).value ='數量  '
        s1.cell(2,4).value ='單位  '
        s1.cell(2,5).value ='單價  '
        s1.cell(2,6).value ='小計  '
        y=3
        
        for i in range(len(temp)):
            s1.cell(y,1).value =temp[i][0]
            s1.cell(y,2).value =temp[i][1]
            s1.cell(y,3).value =temp[i][2]
            s1.cell(y,4).value =temp[i][3]
            s1.cell(y,5).value =temp[i][4]
            s1.cell(y,6).value =temp[i][5]
            y=y+1

        
        if form_use('form_use.xlsx')==True:
            wb.save('form_use.xlsx')
            test_show_1("execl 已更新資料")
        else:
            test_show_1("execl檔案未關閉 無法更新資料")

        
    

    def wire_print(temp,flag): #p015 #p022
    #02-01-04 #02-03-04
        
        import os
        os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

        import openpyxl
        wb = openpyxl.load_workbook('wire_use.xlsx', data_only=True)

        s1 = wb['sheet1']            # 開啟工作表 1
        s1['A1'].value = '名稱'     # 儲存格 A1 內容為
        s1['B1'].value = '成本'    # 儲存格 A2 內容為 
        s1['C1'].value = '大賣'    # 儲存格 A3 內容為 
        
        wire_temp=[]# 電線名,成本,大賣,標籤
        temp_use=[]
        
        for i in range(len(wire)):
            if(wire[i][2]==1):
                para=int(temp[1])
            elif(wire[i][2]==2):
                para=int(temp[2])
            elif(wire[i][2]==3):
                para=int(temp[3])
            else:
                para=int(temp[4]) 

            lable=find_goods_lableonlya(wire[i][0])
            temp_use.append(wire[i][0])
            temp_use.append(int(int(wire[i][1])*(para)/100))
            temp_use.append(int(int(wire[i][1])*(para+10)/100))
            temp_use.append(lable)
            wire_temp.append(temp_use)
            temp_use=[]
        
            s1.cell(i+2,1).value = wire[i][0]      # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
            s1.cell(i+2,2).value = int(int(wire[i][1])*para/100)    # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
            s1.cell(i+2,3).value = int(int(wire[i][1])*(para+10)/100)     # 儲格 B3 內容 ( row=3, column=2 ) 為 300
        
        #print("test")
        #print(form_use("wire_use.xlsx"))
        if form_use("wire_use.xlsx")==True:
            wb.save('wire_use.xlsx')
            test_show_1("execl 已更新資料")
        else:
            test_show_1("execl檔案未關閉 無法更新資料")

        if flag==8:#p022 #02-03-04
            #test_show_1("this is flag 8")
            for i in range(len(wire_temp)):
                #print(wire_temp[i][0],"lable=",wire_temp[i][3])
                temp_use=json.loads(date_temp.get(str(wire_temp[i][3])))
                #print(temp_use)
                temp_use[1]=wire_temp[i][1]
                temp_use[2]=wire_temp[i][2]
                temp_use[3]=wire_temp[i][2]
                
                temp_use=json.dumps(temp_use)#
                
                #print(str(wire_temp[i][3]))
                #print(path_goods)
                
                update(path_goods,str(wire_temp[i][3]),temp_use)
                temp_use=[]
            
            test_temp.append(66)
            window.destroy()#退出 關閉螢幕 




        if flag==5: #02-01-04
            right_show_wire(0,0,wire_temp) #p016

    
    def check_goods(number,flag):#有相同產品名稱或相同編號####
    
        for i in range(2,goods_a+1):
            d1=date_temp.get(str(i))
            d2=json.loads(d1)
            if(number == d2[0] or number == d2[4] ):
                test_show_1("有相同產品名稱或相同編號")
                #print(list1[0],"  :",end='')
                #number=input()
                flag =1
                print(i)
                break
        return flag

    def left_clean():#清除左邊顯示
        Canvas=tk.Canvas
        cv = Canvas(window,width=300, height=400)#,bg = 'white'
            #           创建一个矩形，坐标为(10,10,110,110)
            #cv.create_rectangle(10,10,110,110)
        cv.place(x=0,y=250)

    def middle_clean():#清除中間顯示
        Canvas=tk.Canvas
        cv = Canvas(window,width=350, height=600)#,bg = 'white'
            #           创建一个矩形，坐标为(10,10,110,110)
            #cv.create_rectangle(10,10,110,110)
        cv.place(x=300,y=0)

    def right_clean():#清除右邊顯示
        Canvas=tk.Canvas
        cv = Canvas(window,width=700, height=700)#,bg = 'white'
            #           创建一个矩形，坐标为(10,10,110,110)
            #cv.create_rectangle(10,10,110,110)fixfix
        cv.place(x=650,y=0)
    
    #0011   #p002
    def find_goods(keyin,flag):#找到 產品  資料  不顯示
        
        list_inquire=[] #收集找到的值
        for i in range(2,goods_a+1):
            d1=date_temp.get(str(i))
            d2=json.loads(d1)
            if(keyin in d2[0] or keyin in d2[4] ):
                list_inquire.append(d2)    
                        
                        

                #print(list_inquire)
        if(len(list_inquire)==0):
            print("找不到資料")   
            test_show_1("找不到資料")
            
        elif(len(list_inquire)==1):
            
            #print("test",list_inquire[0][4],flag)
            find_goods_lable(list_inquire[0][4],flag)
            #01-01-03 #01-02-03 #01-04-03
            #03-02-06
        
        elif(len(list_inquire)>20): 
            test_show_1("請縮小尋找範圍")         
        else:
            #print('@@@@end') #
            #test_show_1(len(list_inquire))
            dict_goods={}
            for i in range(len(list_inquire)):
                dict_goods[i]=(list_inquire[i][0],list_inquire[i][4])
            middle_clean()
            middle_show_b(1,"產品資訊",dict_goods,flag)#01-01-03a #p005
            #01-04-03a #03-02-06a
    #0012  #p003
    def find_goods_lable(keyin,flag):#找到 產品標籤 資料  再顯示
        
        list_inquire=[] #收集找到的值
        
        for i in range(2,goods_a+1):
            d1=date_temp.get(str(i))
            d2=json.loads(d1)
            if(keyin == d2[4] ):
                #print("label",i)
                list_inquire.append(d2)  
                lable= i
                break
        
        middle_clean()
        #test_show_1(lable)#fixfixfxi
        if flag==1 :#01-01  
            #print("tesst",list_inquire[0],)
            middle_show_a(1,"產品資訊",list_inquire[0],flag)#01-01-04
        if flag==4:
            
            middle_show_a(2,"產品資訊",list_inquire[0],flag)#01-04-04
        if flag==2:
            #test_show_1("flag")
            middle_show_b(2,"修改產品資料",list_inquire[0],flag)#01-02-04
        if flag==10:#03-02 #p031 
            print(list_inquire[0],flag) 
            keyin_entry_form(3,"輸入數量",10,list_inquire[0]) #03-02-07
        if flag==11:#03-02 #p033d
            #print(list_inquire[0],flag) #
            keyin_entry_form_fix(3,"輸入數量(修改)",11,list_inquire[0]) 
        #left_clean()
    #0016  
    def find_goods_lableonly_b(keyin):#編號 找到 產品標籤
        
        #list_inquire=[] #收集找到的值
        
        for i in range(2,goods_a+1):
            d1=date_temp.get(str(i))
            d2=json.loads(d1)
            if(keyin == d2[0] ):
                #list_inquire.append(d2)  
                
                lable= d2
                break
        return lable
    
    #0013
    def find_goods_lableonly(keyin):#編號 找到 產品標籤
        
        #list_inquire=[] #收集找到的值
        
        for i in range(2,goods_a+1):
            d1=date_temp.get(str(i))
            d2=json.loads(d1)
            if(keyin == d2[4] ):
                #list_inquire.append(d2)  
                lable= i
                break
        
        return lable

    #0015
    def find_goods_lableonlya(keyin):#產品名 找到 產品標籤
        
        #list_inquire=[] #收集找到的值
        
        for i in range(2,goods_a+1):
            d1=date_temp.get(str(i))
            d2=json.loads(d1)
            if(keyin == d2[0] ):
                #list_inquire.append(d2)  
                lable= i
                break
        
        return lable


        #03-02 #p029
    def keyin_entry_form(mode,title,flag,temp):# 模式   標頭名稱
        #if isValidDate(entry1.get()) != False:
        use=[]
        def onok():
            if entry1.get()!= "":
                if flag==10: #03-02-02
                    if mode==1:
                        if isValidDate(entry1.get()) == True:
                        
                            head_temp.append(entry1.get())
                            
                            serial="N"+str(entry1.get())+str(serial_net)
                            head_temp.append(serial)
                            #test_show_1(head_temp) #03-02-03
                            middle_show_b(1,"廠商 客戶 名稱",dict_business_list,10)#p030
                    if mode==2:
                        find_goods(entry1.get(),flag)#031
                        test_show_1(entry1.get()) #03-02-05
                    if mode==3:
                        #keyin=entry1.get()
                        #print("3",entry1.get())
                        
                        one_form.append(float(entry1.get()))
                        one_form.append(temp[5])
                        keyin_entry_form(4,"輸入單價",10,temp) #03-02-08
                    if mode==4:
                        
                        
                        
                        one_form.append(float(entry1.get()))
                        
                        one_form.append(float(one_form[1])*float(one_form[3]))
                        #print("4",one_form)
                        #print(head_temp)
                        #print(serial)
                        #use.append(entry1.get())
                        #use.append(temp[5])
                        keyin_entry_form(5,"total",10,temp) #03-02-09
                    

        
        
        def save():#p032a #03-02-10
            test_show_1("save")               
            #print(head_temp)
            #print(one_form)
            test=[0,0,0,0,0]
            for i in range(len(test)):
                test[i]=one_form[i]
            

            total_form.append(test)
            
            left_clean()
            right_clean()
            print(head_temp)
            right_show(head_temp[0],head_temp[1],head_temp[2],total_form,flag)
            #03-02-11
        def notsave():
            for i in range(0,5):
                    del one_form[0]
                #print("new",one_form)
            left_clean()
            keyin_entry_form(2,"輸入產品名",10,0)
            
            
            
        


        #left_clean()
        title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
        title.place(x=0,y=250)
        if mode!=5:   
            entry1 = tk.Entry(window, width = 20,bd=5,font = ('Arial', 16))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
            entry1.place(x=0,y=280)
            button = tk.Button(window, text="ok" ,fg="#8B008B", bg="#7AFEC6", width=10,font = ('Arial', 16),
            command=onok)
            button.place(x=60,y=310)
        
        
            
        if mode==1:#只顯示不修改
            
            title = tk.Label(window, text='日期:' ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=360)
            title = tk.Label(window, text='客戶:' ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=395)
            
        elif mode==2:#mode==2
            title = tk.Label(window, text='日期:'+head_temp[0] ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=360)
            title = tk.Label(window, text='客戶:' +head_temp[2] ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=395)
            
            title = tk.Label(window, text='產品名:' ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=430)
            
        elif mode==3:#p031
            
            #title = tk.Label(window, text='日期:'+head_temp[0] ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            #title.place(x=0,y=360)
            #title = tk.Label(window, text='客戶:' +head_temp[1] ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            #title.place(x=0,y=395)
            
            title = tk.Label(window, text='產品名:'+"("+temp[0]+")" ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=430)
            
            one_form.append(temp[0])#產品名
            title = tk.Label(window, text='數量:' ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=465)
            
        elif mode==4:
            
            title = tk.Label(window, text='數量:' +str(one_form[1])+one_form[2],fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            #print(use)
            title.place(x=0,y=465)
            for i in range(len(customer)):
                if head_temp[2]==customer[i]:
                    flag_a=1#客戶
                    
                    #test_show_1(flag_a)
                    break
                else:
                    
                    flag_a=2
            #print("te",head_temp[2])#
            #print("flag_a=",flag_a)
            if flag_a==1:
                title = tk.Label(window, text='單價:'+"("+temp[2]+")" ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
                title.place(x=0,y=500)
            else:
                title = tk.Label(window, text='單價:'+"("+temp[1]+")" ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
                title.place(x=0,y=500)
            
        elif mode==5:
            title = tk.Label(window, text='單價:'+str(one_form[3]) ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=500)
            title = tk.Label(window, text='小計:' +str(one_form[4]),fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=535)
            
            
            button_l = tk.Button(window, text="儲存單筆 " ,fg="#8B008B", bg="#7AFEC6", width=10,font = ('Arial', 16),
            command=save)
            button_l.place(x=0,y=570)
            button_r = tk.Button(window, text="不儲存" ,fg="#8B008B", bg="#7AFEC6", width=10,font = ('Arial', 16),
            command=notsave)
            button_r.place(x=150,y=570)

    #03-02 #p033c
    def keyin_entry_form_fix(mode,title,flag,temp):# 模式   標頭名稱
        #if isValidDate(entry1.get()) != False:
        

        def onok():
            if entry1.get()!= "":
                if flag==11:
                    
                    if mode==2:
                        #print(entry1.get(),flag)
                        find_goods(entry1.get(),flag)#
                        #test_show_1(entry1.get())
                    if mode==3:
                        #keyin=entry1.get()
                        #print("3",entry1.get())
                        #print("3 pass")
                        one_form.append(float(entry1.get()))
                        one_form.append(temp[5])
                        keyin_entry_form_fix(4,"輸入單價(修改)",11,temp)
                    if mode==4:
                        
                        
                        
                        one_form.append(float(entry1.get()))
                        
                        one_form.append(float(one_form[1])*float(one_form[3]))
                        
                        keyin_entry_form_fix(5,"total",11,temp)
                    

        
        
        def save():#p032a
            test_show_1("save")               
            #print(head_temp)
            #print(one_form)
            #print(form_choose)
            test=[0,0,0,0,0]
            for i in range(len(test)):
                test[i]=one_form[i]
            
            total_form[form_choose[0]]=test
            #total_form.append(test)
            
            left_clean()
            right_clean()
            right_show(head_temp[0],head_temp[1],head_temp[2],total_form,10)
            
        def notsave():
            print("notsave")
            
            
            
        


        #left_clean()
        title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
        title.place(x=0,y=250)
        if mode!=5:   
            entry1 = tk.Entry(window, width = 20,bd=5,font = ('Arial', 16))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
            entry1.place(x=0,y=280)
            button = tk.Button(window, text="ok" ,fg="#8B008B", bg="#7AFEC6", width=10,font = ('Arial', 16),
            command=onok)
            button.place(x=60,y=310)
        
        
            
        
            
        if mode==2:#mode==2
            title = tk.Label(window, text='日期:'+head_temp[0] ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=360)
            title = tk.Label(window, text='客戶:' +head_temp[2] ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=395)
            
            title = tk.Label(window, text='產品名:' ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=430)
            
        elif mode==3:

            
            title = tk.Label(window, text='產品名:'+"("+temp[0]+")" ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=430)
            
            one_form.append(temp[0])#產品名
            #print("test",one_form)
            title = tk.Label(window, text='數量:' ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=465)
            
        elif mode==4:
            
            title = tk.Label(window, text='數量:' +str(one_form[1])+one_form[2],fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            #print(use)
            title.place(x=0,y=465)
            for i in range(len(customer)):
                if head_temp[2]==customer[i]:
                    flag_a=1#客戶
                    break
                else:
                    flag_a=2
            if flag_a==1:
                title = tk.Label(window, text='單價:'+"("+temp[2]+")" ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
                title.place(x=0,y=500)
            else:
                title = tk.Label(window, text='單價:'+"("+temp[1]+")" ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
                title.place(x=0,y=500)
            
        elif mode==5:
            title = tk.Label(window, text='單價:'+str(one_form[3]) ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=500)
            title = tk.Label(window, text='小計:' +str(one_form[4]),fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=535)
            
            
            button_l = tk.Button(window, text="儲存單筆 " ,fg="#8B008B", bg="#7AFEC6", width=10,font = ('Arial', 16),
            command=save)
            button_l.place(x=0,y=570)
            button_r = tk.Button(window, text="不儲存" ,fg="#8B008B", bg="#7AFEC6", width=10,font = ('Arial', 16),
            command=notsave)
            button_r.place(x=150,y=570)

    def read_date_temp(lable):#資料庫 讀取
        d1=date_temp.get((lable))
        d2=json.loads(d1)
        return d2
    
    def delete_date_temp(lable):#資料庫 刪除
        del date_temp[str(lable)]
        
    
    def updata_date_temp(lable,temp):#資料庫 (更新或新增)
        #temp=json.dumps(temp)
        dict={(lable):temp}
        date_temp.update(dict)

    #0010 1 只顯示不修改  2 修改 #01-01 #01-02 #p001
    def keyin_entry(mode,title,flag):# 模式   標頭名稱
        
        def onok():
            if entry1.get()!= "":
                if flag==2 or flag==1 or flag==4: #01-02-02 #01-04-02
                    find_goods(entry1.get(),flag) #01-01-02
                elif flag==9:#024
                    if isValidDate(entry1.get()) != False:
                            
                        form_search.append(entry1.get())
                        
                        left_clean()
                        
                        if len(form_search) == 1:
                            keyin_entry(1,"結束時間",9) #03-01-02
                            
                        else:  #03-01-03                          
                            middle_show_b(1,"廠商 客戶 名稱",dict_business_list,9)
                elif flag==12 or flag==13 :#034a 035a
                    if isValidDate(entry1.get()) != False:
                        form_search.append(entry1.get())
                    left_clean()
                        
                    if len(form_search) == 1:
                        keyin_entry(1,"(結束時間)",flag)#
                        #03-03-02   #03-04-02
                    else:                            
                        #03-03-03 #03-04-03
                        form_search_only(form_search,flag)
                        
        left_clean()
        

        if mode==1:#只顯示不修改
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=250)
            entry1 = tk.Entry(window, width = 20,bd=5,font = ('Arial', 16))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
            entry1.place(x=0,y=280)
            button = tk.Button(window, text="ok" ,fg="#8B008B", bg="#7AFEC6", width=10,font = ('Arial', 16),
            command=onok)
            button.place(x=60,y=310)

        else:
            #left_clean()
            print("enf")

    ####0001a
    def left_show_a(mode,title,temp):#模式 0 不顯 1 顯示,表頭,資料
        #print("left show=",date_temp.get("a"))
        
        
        mode=mode
        title=title
        temp=temp
        #main_read()
        #left_clean()
        middle_clean()
        right_clean()
        
        if mode==0:
            def Selection ():
                test_show_1(var.get())  #change
                if var.get()==1:
                    #keyin_entry(1,"產品名稱",1)#01-00
                    left_show_a(1,"產品查詢",A1)
                    
                elif var.get()==2: #p013
                    #keyin_entry(1,"產品名稱",1)#01-00
                    left_show_a(2,"電線查詢",B1)
                elif var.get()==3: #p023
                    #keyin_entry(1,"產品名稱",1)#01-00
                    left_show_a(3,"表單查詢",C1)
                if var.get()==88:
                    window.destroy()#關閉螢幕 break break
            left_start_x=10
            left_start_y=10
            
            var = tk.IntVar()#設置 var 內容
            var.set(0)
            
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=30,font = ('Arial', 11))
            title.place(x=left_start_x,y=left_start_y)
            left_start_y=left_start_y+30

            for val, top in temp.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 11))
                R.place(x=left_start_x,y=left_start_y)
                left_start_y=left_start_y+30
        if mode==1:
            def Selection ():
                middle_clean()
                left_clean()
                
                test_show_1(var.get())  #change
                if var.get()==1:
                    keyin_entry(1,"產品名稱",1)#01-01-01
                    #left_show_a(1,"產品查詢",A11)
                elif var.get()==2:
                    #left_show(1,"修改產品",A1)
                    keyin_entry(1,"修改產品名稱",2) #01-02-01
                elif var.get()==3:
                    temp=["",0,0,0,"","",0,""]  #p011
                    middle_show_b(3,"建立新產品",temp,3)#01-03-01
                elif var.get()==4:
                    keyin_entry(1,"刪除產品名稱",4) #01-04-01
                else:
                    left_show_a(0,"主畫面",tops)
            left_start_x=10
            left_start_y=10
            
            var = tk.IntVar()#設置 var 內容
            var.set(0)
            
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=30,font = ('Arial', 11))
            title.place(x=left_start_x,y=left_start_y)
            left_start_y=left_start_y+30

            for val, top in temp.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 11))
                R.place(x=left_start_x,y=left_start_y)
                left_start_y=left_start_y+30
        if mode==2:  #p013
            test_show_1("2 開始")
            def Selection ():
                middle_clean()
                left_clean()
                
                test_show_1(var.get())  #change
                if var.get()==1:
                    test_show_1(var.get())
                    
                    middle_show_b(4,"電線專用",wire_a,5)#02-01-01
                    
                elif var.get()==2:
                    test_show_1(var.get())
                    
                    keyin_entry2(1,"現有設定檔",wire_a)#p017 #02-02-01
                elif var.get()==3:
                    test_show_1(var.get())
                    middle_show_b(4,"電線專用",wire_a,8)#p022 #02-03-01
                    
                
                else:
                    test_show_1(var.get())
                    test_show_2("heer")
                    left_show_a(0,"主畫面",tops)
            
            left_start_x=10
            left_start_y=10
            
            var = tk.IntVar()#設置 var 內容
            var.set(0)
            
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=30,font = ('Arial', 11))
            title.place(x=left_start_x,y=left_start_y)
            left_start_y=left_start_y+30

            for val, top in temp.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 11))
                R.place(x=left_start_x,y=left_start_y)
                left_start_y=left_start_y+30
        
        if mode==3:  #p023
            test_show_1("3 開始")
            
            def Selection ():
                middle_clean()
                left_clean()
                
                #test_show_1(var.get())  #change
                if var.get()==1:#p023
                
                    test_show_1(var.get())
                    #form_search=[]#用於 日期 店名 收尋用
                    #trans_temp=[]#用於 資料傳輸
                    for i in range(len(form_search)):
                        del form_search[0]
                    for i in range(len(trans_temp)):
                        del trans_temp[0]
                    right_clean()
                    
                    
                    keyin_entry(1,"開始日期",9)#03-01-01
                elif var.get()==2:
                    test_show_1(var.get())
                    
                    
                    for i in range(len(head_temp)):
                        del head_temp[0]
                    for i in range(len(one_form)):
                        del one_form[0]
                    for i in range(len(total_form)):
                        del total_form[0]   
                    for i in range(len(form_choose)):
                        del form_choose[0]    
                    keyin_entry_form(1,"輸入日期",10,0) #03-02-01 #p029    
                elif var.get()==3: #p034  
                    test_show_1(var.get())
                    for i in range(len(head_temp)):
                        del head_temp[0]
                    for i in range(len(one_form)):
                        del one_form[0]
                    for i in range(len(total_form)):
                        del total_form[0]   
                    for i in range(len(form_choose)):
                        del form_choose[0] 
                    right_clean()
                    
                    keyin_entry(1,"(開始日期)",12)#03-03-01
                elif var.get()==4:
                    
                    for i in range(len(form_search)):
                        del form_search[0]
                    for i in range(len(trans_temp)):
                        del trans_temp[0]
                    right_clean()
                    
                    keyin_entry(1,"(開始日期)",13)#03-04-01 #p035a




                    
                
                else:
                    test_show_1(var.get())
                    test_show_2("back home")
                    left_show_a(0,"主畫面",tops)
            
            left_start_x=10
            left_start_y=10
            
            var = tk.IntVar()#設置 var 內容
            var.set(0)
            
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=30,font = ('Arial', 11))
            title.place(x=left_start_x,y=left_start_y)
            left_start_y=left_start_y+30

            for val, top in temp.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 11))
                R.place(x=left_start_x,y=left_start_y)
                left_start_y=left_start_y+30

    
    def test_show_1(temp):###測試顯示用
        title = tk.Label(window, text=temp ,fg="#8B008B", bg="#7AFEC6", width=30)
        title.place(x=0,y=620)###測試顯示用
    
    def test_show_2(temp):###測試顯示用
        title = tk.Label(window, text=temp ,fg="#8B008B", bg="#7AFEC6", width=30)
        title.place(x=0,y=660)###測試顯示用
    
    
    def data_save_or_delete_a(mode,temp): #p036
        
        if mode==1:
            lable=find_goods_lableonly(temp[4])
            temp=json.dumps(temp)    
            
            update(path_goods,lable,temp)##建立 位置/檔名/資料/    
            
            updata_date_temp(str(lable),temp)#只有可以變更  #useuse
            test_show_1("fix  ok")
            #print("psaa")#999
            left_clean()
            middle_clean()
    
    
    
    
    
    #0014
    
    
    
    
    def data_save_or_delete(mode,temp):#01-02-04 #p009
        
        if mode==1:#01-02-10
            lable=find_goods_lableonly(temp[4])
            temp=json.dumps(temp)    
            #print(temp)
            #print(lable)
            update(path_goods,lable,temp)##建立 位置/檔名/資料/    
            
            updata_date_temp(str(lable),temp)#只有可以變更  #useuse
            test_show_1("fix  ok")
            
            left_clean()
            middle_clean()
            right_clean()
            left_show_a(0,"主畫面",tops)# 開始畫面
            #print("test",read_date_temp(lable)) 
            
            
        elif mode==2:#p012  #01-03-06 #02-02-05
            correct=0
            if temp[0] ==" " or temp [4]== "" or temp [5]== "":
                correct=1
            if temp[1] ==0 or temp [2]== 0  or temp [3]== 0 or temp [6]== 0:
                correct=1
            
            if correct ==1 :
                test_show_1("產品資料不完整")
            else:
                test_show_1("存檔")
                
                temp=json.dumps(temp)
                update(path_goods,goods_a+1,temp)##建立 位置/檔名/資料/    
                update(path_goods,'a',goods_a+1)
                
                
                test_temp.append(66)
                window.destroy()#關閉螢幕
        else:#p021 #mode=3 電線用
            if int(temp[0]) !=0 and int(temp[1]) !=0 and int(temp[2]) !=0 and int(temp[3]) !=0 and int(temp[4]) !=0 :
            
                wire_a= json.loads(date_temp.get('wire'))#產品數量_
            
                test_show_1(temp)
                wire_a.append(temp)#####
                wire_a=json.dumps(wire_a)
                update(path_goods,'wire',wire_a)
                test_temp.append(66)
                window.destroy()#關閉螢幕
                
                
                
            else:
                test_show_1("資料不完整 ")   
            
    #0014
    def ask_yes_or_no(mode,temp):#01-04-06              
        lable=find_goods_lableonly(temp[4])
        #print("ttt==",lable)
        
        def sure():
            #print(lable)
            #print("old=",date_temp.get("a"))
            
            #print(goods_a)
            if mode==1:#01-04 use 刪除畫面 
                if int(lable)==int(goods_a):
                    
                    deldate(path_goods+str(lable))
                    update(path_goods,'a',goods_a-1)
                    #update(path_goods,'keep',keep_for_net+1)#for test
                    test_show_1("del_date")
                    #main_read()
                
                else:
                        
                    temp=readdate(path_goods+str(goods_a))
                    #print(temp)
                    update(path_goods,lable,temp)
                    #update(path_goods,'keep',keep_for_net+1)#for test
                    deldate(path_goods+str(goods_a))
                    update(path_goods,'a',goods_a-1)
                    test_show_1("del_date")
                    #print("del_date")
                    #main_read()
            
            test_temp.append(66)
            window.destroy()#退出 關閉螢幕 
            
            
        
        def notsure():
            left_show_a(0,"主畫面",tops)
            left_clean()
            middle_clean()
            
            test_show_1("not delete goods")#01-04-04
        
        
        button = tk.Button(window, text="確定刪除" ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
        command=sure)
        button.place(x=300,y=310)
        button2 = tk.Button(window, text='不刪除' ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
        command=notsure)
        button2.place(x=425,y=310)

        

    #0017 #p017
    def keyin_entry2(mode,title,temp):#   刪除或新增
        
        left_clean()
        def delete():#02-02-07
                
                
            wire_a= json.loads(date_temp.get('wire'))#產品數量_       
            #middle_show(1,"修改畫面",temp,2)
            #middle_show_b(1,"修改畫面",temp,2)
            left_clean()
            if mode==1:#02-02-07
                middle_show_b(4,"現有設定檔",temp,6) 
            elif mode==2:#p019 02-02-09 刪除設定檔 
                #test_show_1(temp)   
                #test_show_2(wire_a)
                for i in range(len(wire_a)):
                    if temp==wire_a[i][0]:
                        keep=i
                        
                del wire_a[keep]        
                #test_show_2(wire_a)
                
                
                if len(wire_a)==0:
                    #test_show_1(i)
                    deldate(path_goods+'wire')
                else:
                    wire_a=json.dumps(wire_a)
                    update(path_goods,'wire',wire_a)   
                test_temp.append(66)
                window.destroy()#關閉螢幕 
                
            
        def save():
            
            if mode==2:
                left_clean()
                left_show_a(0,"主畫面",tops)#
            elif mode==1:
                #test_show_1("here") #p020
                temp=[0,0,0,0,0]
                middle_show_b(3,"新建電線檔",temp,7)#02-02-03
                
            

            #left_clean()
        if mode==1:#02-02-02 一開始用
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=250)
            #entry1 = tk.Entry(window, width = 20, text='y',bd=5,font = ('Arial', 16))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
            #entry1.place(x=0,y=280)
            button = tk.Button(window, text="刪除" ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
            command=delete)
            button.place(x=0,y=310)
            button2 = tk.Button(window, text='新建' ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
            command=save)
            button2.place(x=125,y=310)
        if mode==2: #02-02 選擇刪除用
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=250)
            title = tk.Label(window, text=temp ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=280)
            
            #entry1 = tk.Entry(window, width = 20, text='y',bd=5,font = ('Arial', 16))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
            #entry1.place(x=0,y=280)
            button = tk.Button(window, text="刪除" ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
            command=delete)
            button.place(x=0,y=310)
            button2 = tk.Button(window, text="不須刪除",fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
            command=save)
            button2.place(x=125,y=310)


    
    #0002a 顯示選項 #p004
    def middle_show_a(mode,title,temp,flag):#模式 0 不顯 1 只顯示產品,表頭,資料
        
        
        if mode==1 :#01-01-05 use
            list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
            middle_start_x=300
            middle_start_y=10

            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=30,font = ('Arial', 14))
            title.place(x=middle_start_x,y=middle_start_y)
            middle_start_y=middle_start_y+30
            
            for i in range(len(list1)):
                R=tk.Label(window, text=list1[i]+":"+str(temp[i]),fg="#8B008B", width=30,font = ('Arial', 13))
                R.place(x=middle_start_x,y=middle_start_y)
                middle_start_y=middle_start_y+30
        
        if mode== 2:#01-04-use #p010 
            list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
            middle_start_x=300
            middle_start_y=10

            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=30,font = ('Arial', 14))
            title.place(x=middle_start_x,y=middle_start_y)
            middle_start_y=middle_start_y+30
            
            for i in range(len(list1)):
                R=tk.Label(window, text=list1[i]+":"+str(temp[i]),fg="#8B008B", width=30,font = ('Arial', 13))
                R.place(x=middle_start_x,y=middle_start_y)
                middle_start_y=middle_start_y+30
            ask_yes_or_no(1,temp)#01-04-05
                    #print(temp)
            #test_show_1(temp)
    
    
    
    
    
    
    #0002b #p005
    def middle_show_b(mode,title,temp,flag):#模式 0 不顯 1 只顯示產品,表頭,資料
        list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
        

        def keyin_entry1(title,choose,temp):# 模式   標頭名稱 #p007
            #01-02-06
            left_clean()
            def fix():
                #print("01-02 test",flag)
                #test_show_2(entry1.get())
                if flag==7:#p021 02-02建立新資料用
                    if entry1.get()!= "" and entry1.get().isdigit()==True:
                        temp[choose]=entry1.get()
                        middle_show_b(3,"電線設定檔",temp,flag) 
                    
                        left_clean()
                
                if flag !=3 and flag !=7:
                    if entry1.get()!= "" and entry1.get().isdigit()==True:
                        
                        temp[choose]=entry1.get()
                        middle_show_b(2,"修改畫面",temp,flag) 
                        #01-02-07
                        
                        left_clean()
                        
                if flag==3:
                    #print("flag==3")
                    #test_show_1(choose)
                    if choose==0 or choose==4: #P011 # 01-03 USE
                        test=check_goods(entry1.get(),0)
                        if entry1.get()!= "" and test !=1  :
                            if choose==0:#產品名
                                temp[choose]=entry1.get()
                                temp[7]=entry1.get()
                            else:
                                temp[choose]=entry1.get()#產品編號
                            middle_show_b(3,"修改畫面",temp,flag) #01-03-04
                    elif choose==1 or choose==2 or choose==3 or choose==6:
                        if entry1.get()!= "" and entry1.get().isdigit()==True:
                            temp[choose]=int(entry1.get())#得輸入數字欄位
                            middle_show_b(3,"修改畫面",temp,flag) #01-03-04
                            
                    else:
                        if entry1.get()!= "":
                            temp[choose]=entry1.get()
                            middle_show_b(3,"修改畫面",temp,flag)
                    left_clean()
                    #test_show_1(choose)
            
            def save():
                if flag==7:
                    data_save_or_delete(3,temp)#01-02-04 #p021
                elif flag==2:#elif flag != 3 and flag !=7 and flag !=14:
                    data_save_or_delete(1,temp)#01-02-09 #p009
                    #print("now now",flag)
                elif flag==14:
                    #print("now",flag)
                    data_save_or_delete_a(1,temp) #p036
                else:#flag==3 flag==5
                    data_save_or_delete(2,temp)#01-03-04 #p012 
                    #02-02-05
            #left_clean()
            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", width=20,font = ('Arial', 16))
            title.place(x=0,y=250)
            entry1 = tk.Entry(window, width = 20, text='y',bd=5,font = ('Arial', 16))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
            entry1.place(x=0,y=280)
            button = tk.Button(window, text="fix" ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
            command=fix)
            button.place(x=0,y=310)
            button2 = tk.Button(window, text='save' ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 16),
            command=save)
            button2.place(x=125,y=310)


        if mode==1:#01-01 #p005a
            #print("now flag= ",flag)
            middle_start_x=300
            middle_start_y=10

            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", 
            width=30,font = ('Arial', 13))#,font = ('Arial', 14)
            title.place(x=middle_start_x,y=middle_start_y)
            middle_start_y=middle_start_y+30
            
            def Selection (): #
            
                if flag==1 or flag==2 :
                    number=(temp.get(var.get()))
                    left_clean()
                    find_goods_lable(number[1],flag)#01-01-03b
                    #01-02-03b
                #if flag != 9 and flag != 10:
                    #number=(temp.get(var.get()))
                    #print(number[1])
                        #test_show_1(flag)#01-01
                    
                    #find_goods_lable(number[1],flag)# 
                if flag==10 or flag==11:#p030  flag==10
                    number=(temp.get(var.get()))
                    print(number[1])
                    if len(head_temp)==2:#防止 多項產品的append
                        head_temp.append(number)
                        #test_show_1(head_temp) #p030b
                        middle_clean()
                        keyin_entry_form(2,"輸入產品名",10,0)#03-02-04
                    else:
                        find_goods_lable(number[1],flag)#p030b
                        #03-02-06b
                        #多項產品的append
                    #print("head=",head_temp)
                if flag==9:#else:#flag != 9
                        #test_show_1("here")  #p025                    
                    number=(temp.get(var.get()))
                        
                    form_search.append(number)
                    form_search_only(form_search,flag)#03-01-04
                
                if flag==12 or flag==13:#034c #03-03 #035c
                        #test_show_1("here")     #             
                    number=(temp.get(var.get()))
                    test_show_1(number)    
                    
                    
                    
                    one_list= json.loads(account_temp.get(number[0:11]))
                    one3_list=one_list[2]   
                    
                    head_temp.append(one_list[0])#日期
                    head_temp.append(number[0:11]) #序號
                    head_temp.append(one_list[1])#客戶名
                    for i in range(len(one3_list)):
                        total_form.append(one3_list[i]) #
                        
                    if flag==12:
                        #03-03-05
                        right_show(head_temp[0],head_temp[1],head_temp[2],total_form,10)
                    if flag==13:
                        #03-04-05
                        right_show(head_temp[0],head_temp[1],head_temp[2],total_form,13)
            var = tk.IntVar()#設置 var 內容
            var.set(0)

            for val, top in temp.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 13))
                R.place(x=middle_start_x,y=middle_start_y)
                middle_start_y=middle_start_y+30
        
        elif mode==2:#01-02-08  #01-02-05
            #print(temp)
            middle_start_x=300
            middle_start_y=10

            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", 
            width=30,font = ('Arial', 13))#,font = ('Arial', 14)
            title.place(x=middle_start_x,y=middle_start_y)
            middle_start_y=middle_start_y+30
            
            dict_goods={} 
            for i in range(len(temp)):
                dict_goods[i]=list1[i]+" : "+str(temp[i])
                #print(dict_goods)
            def Selection ():
                #print(temp [var.get()])
                #print(var.get())
                #test_show_1(var.get())
                if(var.get()==1 or var.get()==2 or var.get()==3 or var.get()==6 ):
                        
                    head=list1[var.get()]+str(temp[var.get()])
                    keyin_entry1(head,var.get(),temp)#01-02-06
                    #number=(temp.get(var.get()))
                    #print(number[1])
                    #find_goods_lable(number[1])
            var = tk.IntVar()#設置 var 內容
            var.set(0)

            for val, top in dict_goods.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 13))
                R.place(x=middle_start_x,y=middle_start_y)
                middle_start_y=middle_start_y+30
        elif mode==3 : #p011 #01-03-02
            #test_show_1(flag)
            #print("flag=",flag)
            middle_start_x=300
            middle_start_y=10

            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", 
            width=30,font = ('Arial', 13))#,font = ('Arial', 14)
            title.place(x=middle_start_x,y=middle_start_y)
            middle_start_y=middle_start_y+30
            if flag==3:#01-03-02
                dict_goods={} 
                for i in range(len(temp)):
                    dict_goods[i]=list1[i]+" : "+str(temp[i])
                    #print(dict_goods)
            else:#02-02-04
                heada=["設定年月 ex (11201)","3.5 二芯以下 ex (212)","8mm四芯以下 ex (212)","白扁線設定 ex (212)","電線設定 ex (212)"]
                dict_goods={} 
                for i in range(len(temp)):
                    dict_goods[i]=heada[i]+"--"+str(temp[i])       
            
            def Selection ():
                #print(temp [var.get()])#p011
                #print(var.get())
                
                if flag==3:
                    head=list1[var.get()]+str(temp[var.get()])
                    keyin_entry1(head,var.get(),temp) #01-03-03
                else:
                    head=heada[var.get()]+str(temp[var.get()])
                    keyin_entry1(head,var.get(),temp)  #02-02-04 
                
                
                
                    
                    #number=(temp.get(var.get()))
                    #print(number[1])
                    #find_goods_lable(number[1])
            var = tk.IntVar()#設置 var 內容
            var.set(0)

            for val, top in dict_goods.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 13))
                R.place(x=middle_start_x,y=middle_start_y)
                middle_start_y=middle_start_y+30
                #print("jjjj")
                #middle_clean()           
            
        elif mode==4 : #p014  02-01-02 #02-03-02 #03-02-修改單筆02
            #test_show_1(flag)
            middle_start_x=300
            middle_start_y=10

            title = tk.Label(window, text=title ,fg="#8B008B", bg="#7AFEC6", 
            width=30,font = ('Arial', 13))#,font = ('Arial', 14)
            title.place(x=middle_start_x,y=middle_start_y)
            middle_start_y=middle_start_y+30
            if flag==5 or flag==8 or flag==10:
                dict_goods={} 
                for i in range(len(temp)):
                    dict_goods[i]=str(temp[i])
                    #print(dict_goods)
            elif flag==14:
                dict_goods={} 
                for i in range(len(temp)):
                    dict_goods[i]=str(temp[i][0]) #036b
            else:
                head=["設定年月 ex (11201)","3.5 二芯以下設定 ex (212)","8平方 四芯以下設定 ex (212)","白扁線設定 ex (212)","電線設定 ex (212)"]
                dict_goods={} 
                for i in range(len(temp)):
                    dict_goods[i]=head[i]+"--"+str(temp[i])
                #test_show_1("dsfs")
            
            
            
            
            
            def Selection ():
                if flag==10:
                    #print("flag==10",total_form) #033b
                    for i in range(len(form_choose)):
                        del form_choose[0] 
                    form_choose.append(var.get())
                    print("choose=",form_choose)
                    #print( var.get())
                    for i in range(len(one_form)):
                        del one_form[0]
                    keyin_entry_form_fix(2,"輸入產品名(修改)",11,0)
                if flag==5 or flag==8:#print(temp [var.get()][1])
                    wire_print(temp [var.get()],flag)#p015 #02-01-03 
                    #02-03-03
                elif flag==6:
                    test_show_1(temp [var.get()])
                    keyin_entry2(2,"delete",temp [var.get()][0])#p018
                    #02-02-08
                elif flag==14:
                    k=temp [var.get()][0]
                    test_show_1(temp [var.get()][0]) #036c
                    
                    
                    li=(find_goods_lableonly_b(temp [var.get()][0]))
                    middle_show_b(2,"修改產品資料",li,flag)
                    #print(k,li)



            var = tk.IntVar()#設置 var 內容
            var.set(0)

            for val, top in dict_goods.items():
                R=tk.Radiobutton (window,text=top,indicatoron=0,variable=var, value=val,
                command=Selection,bg='yellow',width=30,font = ('Arial', 13))
                R.place(x=middle_start_x,y=middle_start_y)
                middle_start_y=middle_start_y+30
                #print("jjjj")
                #middle_clean()     
                
                
            
        
    #0003a
    def right_show_wire(mode,titel,temp):  #p016 
        right_clean()
        right_start=650
        labela = tk.Label(window, text = '產   品   名   稱  ',font = ('Arial', 14))
        labelb = tk.Label(window, text = '成  本  ',font = ('Arial', 14))
        #labelc = tk.Label(window, text = '成  本  ',font = ('Arial', 14))
        labeld = tk.Label(window, text = '賣   價',font = ('Arial', 14))
        #labele = tk.Label(window, text = '賣   價  ',font = ('Arial', 14))
        labelf = tk.Label(window, text = '備   註  ',font = ('Arial', 14))
            #labela.pack(side='left')
        x=right_start
        y=50
        labela.place(x=x,y=y)#產   品   名   稱
        x=x+250
        labelb.place(x=x,y=y)#成本
        x=x+70
        #labelc.place(x=x,y=y)#
        x=x+70
        labeld.place(x=x,y=y)#賣 價
        x=x+70
        #labele.place(x=x,y=y)#
        x=x+70
        labelf.place(x=x,y=y)#備 註
        x=right_start
        y=100

        for i in range(len(temp)):
            label = tk.Label(window, text = temp[i][0],font = ('Arial', 14))
            label.place(x=x,y=y)
            x=x+250
            label = tk.Label(window, text = temp[i][1],font = ('Arial', 14))
            label.place(x=x,y=y)
            x=x+70
            #label = tk.Label(window, text = temp[i][2],font = ('Arial', 14))
            #label.place(x=x,y=y)
            x=x+70
            label = tk.Label(window, text = temp[i][2],font = ('Arial', 14))
            label.place(x=x,y=y)
            x=x+140
            label = tk.Label(window, text = temp[i][3],font = ('Arial', 14))
            label.place(x=x,y=y)
            
            x=right_start
            y=y+25



    #0003 #p032b
    def right_show(date,serial,guest,temp,flag):#日期,序號,廠商名,資料:
        #print(temp)
        #print(flag)
        
        a=1
        right_start=650
        if a==1:
            keep_temp=[]
            
            def delete():# #03-04-07
                test_show_1("delete")
                newname=head_temp[1]+head_temp[2]
                
                deldate(path_form_name+newname)
                deldate(path_form+head_temp[1])
                middle_clean()
                right_clean()
            
            
            def notdelete():   
                left_show_a(0,"主畫面",tops)
                
            def onOK():
                
                print("")
                
            def onnext():##03-02-next
                #keep_temp.append(1)
                test_show_1("next") #p032c
                left_clean()
                #print(one_form)
                for i in range(0,5):
                    del one_form[0]
                #print("new",one_form)
                keyin_entry_form(2,"輸入產品名",10,0)
                
            def onfix():##03-02 修改單筆
                #keep_temp.append(2)
                #print(total_form)#033a
                middle_show_b(4,"修改單筆",total_form,10)
                
            def onsave():#p032d
                #keep_temp.append(3)
                last_temp=[]
                last_temp.append(head_temp[0])
                last_temp.append(head_temp[2])
                last_temp.append(total_form)
                serial=head_temp[1]
                
                #print("HEAD temp=",head_temp)
                #print("last serial==",serial)
                #print("serial_net=",serial_net)
                newname=serial+head_temp[2]
                #print("newname=",newname)
                #head.append(second_temp)
                last_temp=json.dumps(last_temp)
                #print("head==" ,head)
                update(path_form,serial,last_temp)#
                update(path_form_name,newname,serial)
                update(path_goods,'s',serial_net+1)
                test_show_1("update")
                right_clean()
                
            def onfixsave(): # #p036
                test_show_1("i am here")
                
                middle_show_b(4,"修改產品成本",total_form,14) #修改單項成本
                #keep_temp.append(4)
                

            label_date=tk.Label(window, text = '日期  ',font = ('Arial', 14))
            label_date_in=tk.Label(window, text = date,font = ('Arial', 14))
            label_serial=tk.Label(window, text = '序號  ',font = ('Arial', 14))
            label_serial_in=tk.Label(window, text = serial,font = ('Arial', 14))
            label_name=tk.Label(window, text = '店家名字  ',font = ('Arial', 14))
            label_name_in=tk.Label(window, text = guest,font = ('Arial', 14))
            x=right_start
            y=0
            label_date.place(x=x,y=y)
            x=x+50
            label_date_in.place(x=x,y=y)
            x=x+100
            label_serial.place(x=x,y=y)
            x=x+50
            label_serial_in.place(x=x,y=y)
            x=x+150
            label_name.place(x=x,y=y)
            x=x+100
            label_name_in.place(x=x,y=y)

            # 標示文字
            labela = tk.Label(window, text = '產   品   名   稱  ',font = ('Arial', 14))
            labelb = tk.Label(window, text = '數 量  ',font = ('Arial', 14))
            labelc = tk.Label(window, text = '單 位  ',font = ('Arial', 14))
            labeld = tk.Label(window, text = '單 價  ',font = ('Arial', 14))
            labele = tk.Label(window, text = '小 計  ',font = ('Arial', 14))
            labelf = tk.Label(window, text = '備 註  ',font = ('Arial', 14))
            #labela.pack(side='left')
            x=right_start
            y=50
            labela.place(x=x,y=y)#產   品   名   稱
            x=x+250
            labelb.place(x=x,y=y)#數 量
            x=x+70
            labelc.place(x=x,y=y)#單 位
            x=x+70
            labeld.place(x=x,y=y)#單 價
            x=x+70
            labele.place(x=x,y=y)#小 計
            x=x+70
            labelf.place(x=x,y=y)#備 註
            x=right_start
            y=100
            total=0
            for i in range(len(temp)):
                label = tk.Label(window, text = temp[i][0],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+250
                label = tk.Label(window, text = temp[i][1],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+70
                label = tk.Label(window, text = temp[i][2],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+70
                label = tk.Label(window, text = temp[i][3],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+70
                label = tk.Label(window, text = temp[i][4],font = ('Arial', 14))
                label.place(x=x,y=y)
                total=total+temp[i][4]
                #total=0
                x=right_start
                y=y+25
            window.update_idletasks()     
            x=right_start
            y=y+25
            label = tk.Label(window, text = "----------------------------------------------------------------------------------------------------------")
            label.place(x=x,y=y)
            x=right_start+480#總計用r
            y=y+25   
            label = tk.Label(window, text = total,font = ('Arial', 14))
            label.place(x=x,y=y)
            #print("check",flag) #
            if flag == 10 or flag ==11 or flag ==12:
                blank = 30
                idicent = 170
                button_next = tk.Button(window, text = "下一筆資料", command = onnext,bg='yellow')
                
                #button_next.place(x=idicent*0,y=window.winfo_width()-30)
                button_next.place(x=window.winfo_width()-right_start,y=window.winfo_height()-30)
                
                button_fix = tk.Button(window, text = "修改單筆資料", command = onfix,bg='gray')
                button_fix.place(x=window.winfo_width()-right_start+(idicent*1),y=window.winfo_height()-30)
                button_save = tk.Button(window, text = "儲存表單", command = onsave,bg='yellow')
                button_save.place(x=window.winfo_width()-right_start+(idicent*2),y=window.winfo_height()-30)
                button_fix_save = tk.Button(window, text = "修改成本", command = onfixsave,bg='gray')
                button_fix_save.place(x=window.winfo_width()-right_start+(idicent*3),y=window.winfo_height()-30)
                entry = tk.Entry(window, width = 20,bd=5,font = ('Arial', 18))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
                entry.place(x=0,y=window.winfo_width()-100)
            if flag==13: #03-04-06   #
                button_next = tk.Button(window, text = "刪 除 ", command = delete,bg='yellow',font = ('Arial', 16))

                button_next.place(x=850,y=650)
                button_save = tk.Button(window, text = "不用刪除", command = notdelete,bg='yellow',font = ('Arial', 16))
                button_save.place(x=1000,y=650)
                #print("x width==",window.winfo_width())
                #print("y height==",window.winfo_height())
        else:
            right_clean()
    #0003b #p027
    def right_show_b(temp,name,last_toatl):#日期,序號,廠商名,資料:
        #test_show_1(len(temp))
        
        
        def show(start,end,x,y):#p028
            
            def next_page(): #03-01-06
            
                #test_show_1(page_end)#
                right_clean()
                #test_show_2(page_end[1])
                test_start=page_end[1]
                if test_start==end_line:
                    test_start=page_end[0]-1
                

                if test_start+22 > end_line:
                    show(test_start+1,end_line,right_start+100,100)
                else:
                    show(test_start+1,test_start+23,right_start+100,100)
        
            def back_page():#03-01-07
                #test_show_1(page_end)
                right_clean()
                
                
                test_start=page_end[1]
                if test_start <=end_line:
                    test_start=0
                    test_start=page_end[0]-1
                    if test_start < 0:
                        test_start=0
                #test_show_2(test_start) #
                
                if test_start < 22:
                
                    show(0,end_line,right_start+100,100)
                elif test_start == 22:
                    show(0,22,right_start+100,100)
                
                    
                else:
                    show(test_start-22,test_start,right_start+100,100)

            label_name=tk.Label(window, text = '客戶名稱  ',font = ('Arial', 14))
            label_name_in=tk.Label(window, text = name,font = ('Arial', 14))#
            label_name.place(x=650,y=0)
            label_name_in.place(x=850,y=0)
            labela = tk.Label(window, text = '日 期  ',font = ('Arial', 14))
            labelb = tk.Label(window, text = '產   品   名   稱  ',font = ('Arial', 14))
            labelc = tk.Label(window, text = '數 量  ',font = ('Arial', 14))
            labeld = tk.Label(window, text = '單 位  ',font = ('Arial', 14))
            labele = tk.Label(window, text = '單 價  ',font = ('Arial', 14))
            labelf = tk.Label(window, text = '小 計  ',font = ('Arial', 14))
            labela.place(x=650,y=50)#日 期
            labelb.place(x=750,y=50)#產   品   名   稱
            labelc.place(x=1000,y=50)#數 量
            labeld.place(x=1070,y=50)#單 位
            labele.place(x=1140,y=50)#單 價
            labelf.place(x=1210,y=50)#小 計
            button = tk.Button(window, text="next" ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 14),
            command=next_page)
            button.place(x=950,y=10)
            button2 = tk.Button(window, text='back' ,fg="#8B008B", bg="#7AFEC6", width=9,font = ('Arial', 14),
            command=back_page)
            button2.place(x=1120,y=10)
            
            if end > end_line:
                end=end_line
            
            
            for i  in range(start,end):
                label = tk.Label(window, text = temp[i][0],font = ('Arial', 14))
                label.place(x=x-100,y=y)
                
                label = tk.Label(window, text = temp[i][1],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+250
                label = tk.Label(window, text = temp[i][2],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+70
                label = tk.Label(window, text = temp[i][3],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+70
                label = tk.Label(window, text = temp[i][4],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=x+70
                label = tk.Label(window, text = temp[i][5],font = ('Arial', 14))
                label.place(x=x,y=y)
                x=right_start+100#修改地
                y=y+25
                
            page_end=[]
            page_end.append(start)
            page_end.append(end)
            test_show_2(page_end)
            #副程式結束
        
        right_start=650
        start=0
        end_line=len(temp)
        if len(temp)<=22:
            show(start,end_line,right_start+100,100)
        else:
            show(start,22,right_start+100,100)
        
        
        
        
    
    
    end=1
    if end == 1:
        keep=[]
        keep_temp=[]
        """tops={1:"產品查詢 ", 2 :"電線查詢 ", 3:"表單查詢 ",4 :"貨品進出",  5 :"客戶/廠商 設定",88 :"結束程式"}
        A1={1:"產品查詢 ", 2 :"修改產品資料 ", 3:"建立產品資料 ",4 :"刪除產品資料",  88 :"結束產品查詢",66 :""}
        A11={79:"要存檔 ", 80 :"不要存檔 ",66 :"",67 :" ",68:" ",69 :" "}
        A12={79:"確定刪除 ", 80 :"不要刪除 ",66 :"",67 :" ",68:" ",69 :" "}
        B1={1:"查詢電線價錢與列印 ", 2 :"電線設定檔 ", 3:"更改電線成本 ",  88:"結束電線查詢",66 :"",67 :" "}
        B11={79:"建立電線設定檔",80 :"刪除電線設定檔 ",66 :"",67 :" ",68:" ",69 :" "}
        C1={1:"僅供表單查詢 ", 2 :"增加 /刪除表單", 88:"結束表單查詢",66 :"",67 :" ",68:" "}
        C11={1:"增加新表單", 2 :"修改/ 刪除單日表單 ",66 :"",67 :" ",68:" ",69 :" "}
        C12={1:"修改表單", 2 :"刪除單日表單 ",66 :"結束表單查詢",67 :" ",68:" ",69 :" "}
        D1={1:"全品項查詢", 2 :"單一品項查詢 ",  88:"結束貨品進出",66 :"",67 :" ",68:" "}
        E1={1:"編輯 廠商/客戶 資料  ", 2 :"增加 廠商/客戶 資料", 3:"刪除 廠商/客戶 資料 ",88:"結束貨品進出",66 :"",67 :" "}    
        list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
        wire=[["2平方二芯電纜線",1089,1],
            ['2平方三芯電纜線',1425,1],
            ['3.5平方二芯電纜線',1520,1],
            ['3.5平方三芯電纜線',2079,2],
            ['5.5平方二芯電纜線',2235,2],
            ['5.5平方三芯電纜線',3076,2],
            ['8平方二芯電纜線',3106,2],
            ['8平方三芯電纜線',4313,2],
            ['14平方三芯電纜線',7436,3],
            ['1.6二芯白扁線',700,3],
            ['2.0二芯白扁線',1000,3],
            ['1.6單芯電線',232,4],
            ['2.0單芯電線',350,4],
            ['3.5平方電線',459,4],
            ['5.5平方電線',708,4],
            ['8平方電線',1009,4]]
        date="20221216"
        serial="123456789"
        guest="123456789"
        temp=[["電磁開關MSO-P40(15HP)",1,"個",100,1000],["789456987555",20,"個",15,300]]"""
        

        #testtest
        
        keep_temp=[]
        date_temp=[]
        path_goods='tbuiness/'#產品與電線目錄
        date_temp=(readdate(path_goods))#產品與電線用
        #print("read read")
        business_b= int(date_temp.get('b'))#客戶數量
        factory_f= int(date_temp.get('f'))#廠商數量
        goods_a= int(date_temp.get('a'))#產品數量
        wire_a= json.loads(date_temp.get('wire'))#產品數量_
        #keep_for_net=readdate(path_goods+"keep")
        #print("net==",keep_for_net)
        #keep_for_program=int(date_temp.get('keep'))#1229use
        #print("program==",keep_for_program)
            ####拿到所有廠商客戶名子
        business_list=[]
        for i in range(1001,int(factory_f)+1):
            #print(json.loads(date_temp.get(str(i)))[1])
            business_list.append(json.loads(date_temp.get(str(i)))[1])
        for i in range(1101,int(business_b)+1):
            #print(json.loads(date_temp.get(str(i)))[1])
            business_list.append(json.loads(date_temp.get(str(i)))[1])
            #########放在business_list
        #print(business_list)
        dict_business_list={}
        for i in range(len(business_list)):
            dict_business_list[i]=business_list[i]
        #print(goods_a)
        
        customer=[]#拿到所有 客戶名子
        for i in range(1101,int(business_b)+1):
            #print(json.loads(date_temp.get(str(i)))[1])
            customer.append(json.loads(date_temp.get(str(i)))[1])
        
        #testtest
        
        #print(customer)
        #right_show(1,date,serial,guest,temp)#日期,序號,廠商名,資料:
        #right_show(0,date,serial,guest,temp)#日期,序號,廠商名,資料:
        left_show_a(0,"主畫面",tops)# 開始畫面
        #print(" test==",test_temp)
        #keyin_entry(1,"test")
        #keyin_entry(0,"test")
       #middle_show(2,"選擇廠商 /客戶",dict_business_list)
        #middle_show(0,"產品資訊",["1234",10,10,10,"k88","各",10,"1234"])
        
        #print("choose==",keep_temp)
        
        #end_temp=[]
    
    
    
    window.mainloop()

def readdate(dateset):#讀取(位置+標籤)####
    ref=db.reference(dateset)
    #print(ref.get())                 
    return(ref.get())

def update(dateset,datenames,date):##建立 位置/標籤/資料/
        ref=db.reference(dateset)
        ref.update({datenames:date})

def deldate(dateset):#刪除位置(位置+標籤)
        ref=db.reference(dateset)
        ref.set({})

test_temp=[]
date_temp=[]
keep=[]
form_choose=[]#表單 內 修改的選項

one_form=[]#表單內 單筆資料
total_form=[]#表單內 所有的單筆資料
head_temp=[]#給表單前頭用  #(日期 序號 廠商名)
keep_temp=[]#給表單傳輸資料
serial=""#給表單序號
form_search=[]#用於 日期 店名 收尋用 ( 開始日期  結束日期 廠商名)
trans_temp=[]#用於 資料傳輸
page_end=[]
tops={1:"產品查詢 ", 2 :"電線查詢 ", 3:"表單查詢 ",4 :"貨品進出",  5 :"客戶/廠商 設定",88 :"結束程式"}
A1={1:"產品查詢 ", 2 :"修改產品資料 ", 3:"建立產品資料 ",4 :"刪除產品資料",  88 :"結束產品查詢",66 :""}
A11={79:"要存檔 ", 80 :"不要存檔 ",66 :"",67 :" ",68:" ",69 :" "}
A12={79:"確定刪除 ", 80 :"不要刪除 ",66 :"",67 :" ",68:" ",69 :" "}
B1={1:"查詢電線價錢與列印 ", 2 :"電線設定檔 ", 3:"更改電線成本 ",  88:"結束電線查詢",66 :"",67 :" "}
B11={79:"建立電線設定檔",80 :"刪除電線設定檔 ",66 :"",67 :" ",68:" ",69 :" "}
C1={1:"僅供表單查詢(輸出excel) ", 2 :"增加 表單", 3:"修改 表單", 4:"刪除 表單",88:"結束表單查詢",67 :" "}
C11={1:"增加新表單", 2 :"修改/ 刪除單日表單 ",3 :"刪除單日表單",67 :" ",68:" ",69 :" "}
C12={1:"修改表單", 2 :"刪除單日表單 ",66 :"結束表單查詢",67 :" ",68:" ",69 :" "}
D1={1:"全品項查詢", 2 :"單一品項查詢 ",  88:"結束貨品進出",66 :"",67 :" ",68:" "}
E1={1:"編輯 廠商/客戶 資料  ", 2 :"增加 廠商/客戶 資料", 3:"刪除 廠商/客戶 資料 ",88:"結束貨品進出",66 :"",67 :" "}    
list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
wire=[["2平方二芯電纜線",1089,1],
        ['2平方三芯電纜線',1425,1],
            ['3.5平方二芯電纜線',1520,1],
            ['3.5平方三芯電纜線',2079,2],
            ['5.5平方二芯電纜線',2235,2],
            ['5.5平方三芯電纜線',3076,2],
            ['8平方二芯電纜線',3106,2],
            ['8平方三芯電纜線',4313,2],
            ['14平方三芯電纜線',7436,3],
            ['1.6二芯白扁線',700,3],
            ['2.0二芯白扁線',1000,3],
            ['1.6單芯電線',232,4],
            ['2.0單芯電線',350,4],
            ['3.5平方電線',459,4],
            ['5.5平方電線',708,4],
            ['8平方電線',1009,4]]
        
        
        
path_goods='tbuiness/'#產品與電線目錄
date_temp=(readdate(path_goods))#產品與電線用
        
business_b= int(date_temp.get('b'))#客戶數量
factory_f= int(date_temp.get('f'))#廠商數量
goods_a= int(date_temp.get('a'))#產品數量
wire_a= json.loads(date_temp.get('wire'))#產品數量_
keep_for_net=readdate(path_goods+"keep")
account_temp=[]
path_form='account/'
account_temp=(readdate(path_form))#表單資料
path_form_name='inguire/'
account_name_temp=(readdate(path_form_name))#表單資料+name


serial_net=readdate(path_goods+"s")
if serial_net==99:
    serial_net=11
    update(path_goods,'s',serial_net)
            ####拿到所有廠商客戶名子
business_list=[]
for i in range(1001,int(factory_f)+1):
            #print(json.loads(date_temp.get(str(i)))[1])
    business_list.append(json.loads(date_temp.get(str(i)))[1])
for i in range(1101,int(business_b)+1):
            #print(json.loads(date_temp.get(str(i)))[1])
    business_list.append(json.loads(date_temp.get(str(i)))[1])
            #########放在business_list
        #print(business_list)
dict_business_list={}
for i in range(len(business_list)):
    dict_business_list[i]=business_list[i]

        
customer=[]#拿到所有 客戶名子
for i in range(1101,int(business_b)+1):
            #print(json.loads(date_temp.get(str(i)))[1])
    customer.append(json.loads(date_temp.get(str(i)))[1])

def date_get(name):
    temp=date_temp.get(str(name))
    return temp

def date_increase(name,value):
    date_temp.setdefault(str(name),value)

def date_delete(name):
    del date_temp[str(name)]

def date_change(name,value):
    date_temp[str(name)]=value


    

#test1={"a":1,"b":2,"c":3}
#print(test1.get("b"))
#test1["a"]=30

#print(test1)






show_allform_window()
#print("222=",test_temp[0])
if len(test_temp) != 0:
    while test_temp[0]==66:
        test_temp=[]
        show_allform_window()

