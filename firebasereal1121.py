# 引用必要套件
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
from firebase_admin import db
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import json
import datetime
from datetime import date


# 引用私密金鑰
# path/to/serviceAccount.json 請用自己存放的路徑
cred = credentials.Certificate('/Users/home/Desktop/python/outdata-6457c-firebase-adminsdk-2miud-7b496d14a7 (1).json')

# 初始化firebase，注意不能重複初始化
firebase_admin.initialize_app(cred,{
	'databaseURL':'https://outdata-6457c.firebaseio.com'})


def readdate(dateset):#讀取
    ref=db.reference(dateset)
    #print(ref.get())                 
    return(ref.get())

def update(dateset,datenames,date):##建立 位置/檔名/資料/
    ref=db.reference(dateset)
    ref.update({datenames:date})

def deldate(dateset):#刪除位置
    ref=db.reference(dateset)
    ref.set({})

def show_goods(Atemp):
    for i in range(len(Atemp)):
        if i==1:
            print(list1[i],"-----",Atemp[i]
                  ,"______","加稅價","-----","{:.2f}".format(float(Atemp[i])*1.05))
        else:
            print(list1[i],"-----",Atemp[i])

def firebase_get(a):
    date_temp = json.loads(readdate(a))
    #print(date_temp)
    return(date_temp)

def show_number(a):
    print()
    for i in range(2,goods_a):
        d1=date_temp.get(str(i))
        d2=json.loads(d1)
        if(a==d2[4]):
            show_goods(d2)
             

def geteveryday(begin_date,end_date):
    date_list = []
    begin_date = datetime.datetime.strptime(begin_date, "%Y%m%d")
    end_date = datetime.datetime.strptime(end_date,"%Y%m%d")
    while begin_date <= end_date:
        date_str = begin_date.strftime("%Y%m%d")
        date_list.append(date_str)
        begin_date += datetime.timedelta(days=1)
    return date_list

def isValidDate(datestr):
     try:
         date.fromisoformat(datestr)
     except:
         return False
     else:
         return True
def show_form(one_list,count,y,x,last_total):
        y=y
        x=x
        last_total=last_total
        one3_list=one_list[2]
        print("日期  :",(one_list[0]))
     
        s1.cell(y,x).value =one_list[0]
        x=2      
        if count==0:
            print("客戶  :",(one_list[1]))
            s1.cell(1,1).value ="客戶名稱  " 
            s1.cell(1,2).value =one_list[1]
            s1.cell(2,1).value ="日　期  "            
            s1.cell(2,2).value ="產　品　名　稱"           
            s1.cell(2,3).value ="數　量"          
            s1.cell(2,4).value ="單　位"         
            s1.cell(2,5).value ="單　價"          
            s1.cell(2,6).value ="小　計"

        total=0
        for i in range(len(one3_list)):
            total=total+one3_list[i-1][4]
            print("品名:",'{:15}'.format(one3_list[i-1][0]) ,'%-5s' %one3_list[i-1][1],one3_list[i-1][2],
              '單價','%-5s' %one3_list[i-1][3],"小計:",'%-5s' %one3_list[i-1][4])
            
            s1.cell(y,x).value =one3_list[i-1][0]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][1]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][2]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][3]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][4]
            y=y+1
            x=2
        last_total=last_total+total
        s1.cell(y,1).value="總計"
        s1.cell(y,6).value=last_total

        print ("總計: ",total)
        print("-------------------------------------------------------")
        wb.save('form_use.xlsx')
        return total,y,last_total
    
#IN_temp=readdate('tbuiness/6')

#IN_temp = json.loads(readdate('tbuiness/7'))#將jason轉成array
#str1 = json.dumps(j)#dict包含list轉JSON字串


#show_goods(IN_temp)

#for hh in date_temp:#拿到字典的key {key,value}
#goods_a=int(readdate('tbuiness/a'))
#print (goods_a)

date_temp=[]
test2='tbuiness/'
date_temp=(readdate(test2))#產品與電線用
business_b= int(date_temp.get('b'))#產品數量
factory_f= int(date_temp.get('f'))#產品數量
####拿到所有廠商客戶名子
business_list=[]
for i in range(1001,int(factory_f)+1):
    #print(json.loads(date_temp.get(str(i)))[1])
    business_list.append(json.loads(date_temp.get(str(i)))[1])
for i in range(1101,int(business_b)+1):
    #print(json.loads(date_temp.get(str(i)))[1])
    business_list.append(json.loads(date_temp.get(str(i)))[1])
#########放在business_list





print(" 1 :產品查詢\n","2 :電線查詢\n","3 :表單查詢\n","4 :貨品進出\n")
choose=int(input())


if(choose==1):
    total=0
    IN_tempp=[]
    list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
    test2=[]
    
    
    goods_a= int(date_temp.get('a'))#產品數量
    correct=0          
    list_inquire=[]
    number=input("產品查詢名稱或編號 :   ")
    for i in range(2,goods_a):
        if(correct!=1):
            d1=date_temp.get(str(i))
            d2=json.loads(d1)
            if(number in d2[0] or number in d2[4] ):
            #list_inquire.append(i)
                list_inquire.append(d2)
                correct=0

    if(len(list_inquire)==0):
        print("找不到資料")   
    elif(len(list_inquire)==1):
        print(number) 
        show_number(number)
    else:
        for i in range(len(list_inquire)):
            print(i+1,"  ",list_inquire[i-1][0],"(",list_inquire[i-1][4],")")

    
        number=int(input("選項號碼   :"))

        while(number > len(list_inquire)):
            number=int(input("選項號碼   :"))

        show_number(list_inquire[number-2][4])    

        print('end')
if(choose==2):
    wire_a= json.loads(date_temp.get('wire'))#產品數量
    wire=[["2平方二芯電纜線",1089,1],
      ['2平方三芯電纜線',1425,1],
      ['3.5平方二芯電纜線',1520,1],
      ['3.5平方三芯電纜線',2079,2],
      ['5.5平方二芯電纜線',2235,2],
      ['5.5平方三芯電纜線',3076,2],
      ['8平方二芯電纜線',3106,2],
      ['8平方三芯電纜線',4313,2],
      ['14平方三芯電纜線',7436,3],
      ['1.6二芯白扁線',700,3],
      ['2.0二芯白扁線',1000,3],
      ['1.6單芯電線',232,4],
      ['2.0單芯電線',350,4],
      ['3.5單芯電線',459,4],
      ['5.5單芯電線',708,4],
      [' 8單芯電線',1009,4]
      ]

    for i in range(len(wire_a)):
        print(i+1,"  ",wire_a[i])
    number=int(input("選擇月份項   :"))
    while(number > len(wire_a)):
            number=int(input("選擇月份項   :"))
    
   # title = ["名稱          ","成本","大賣"]
    #print(title)
    import os
    os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

    import openpyxl
    wb = openpyxl.load_workbook('wire_use.xlsx', data_only=True)

    s1 = wb['sheet1']            # 開啟工作表 1
    s1['A1'].value = '名稱'     # 儲存格 A1 內容為 apple
    s1['B1'].value = '成本'    # 儲存格 A2 內容為 orange
    s1['C1'].value = '大賣'    # 儲存格 A3 內容為 banana

    for i in range(len(wire)):
        if(wire[i][2]==1):
            para=int(wire_a[number-1][1])
        elif(wire[i][2]==2):
            para=int(wire_a[number-1][2])
        elif(wire[i][2]==3):
            para=int(wire_a[number-1][3])
        else:
            para=int(wire_a[number-1][4])    

        print("名稱  :",wire[i][0],
              "成本  :",int(int(wire[i][1])*para/100),
              "售價  :",int(int(wire[i][1])*(para+10)/100)
                )
        print()            
        s1.cell(i+2,1).value = wire[i][0]      # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
        s1.cell(i+2,2).value = int(int(wire[i][1])*para/100)    # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
        s1.cell(i+2,3).value = int(int(wire[i][1])*(para+10)/100)     # 儲格 B3 內容 ( row=3, column=2 ) 為 300
    
    
    wb.save('wire_use.xlsx')

if(choose==3):
    
  
    import os
    os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

    import openpyxl
    wb = openpyxl.load_workbook('form_use.xlsx', data_only=True)

    s1 = wb['sheet1']            # 開啟工作表 1

    account_temp=[]
    url='account/'
    account_temp=(readdate(url))#表單資料
    url='inguire/'
    account_name_temp=(readdate(url))#表單資料+name
    

    title=[]
    for  hh in account_name_temp:
        title.append(hh) #表單標籤資料有名字      

    start=input("查詢資料 開始日期 ex 20200413 :  ")
    while start.isdigit()!=True or isValidDate(start)  !=True:
          start=input("查詢資料 開始日期 ex 20200413 :  ")
    end=input("查詢資料 (結束日期) ex 20200413 :  ")
    while end.isdigit()!=True  or isValidDate(end)  !=True:
          end=input("查詢資料 (結束日期) ex 20200413 :  ")
   
    day_choose=geteveryday(start,end)#取得所有日子
    
    second_temp=[]
    for i in range(len(day_choose)):
        for j in range(len(title)):
            if(day_choose[i] in title[j]):
                second_temp.append(title[j])#日期篩選

    #print(second_temp)
    for i in range(len(business_list)):
        print(i+1,"  ",business_list[i])
    number=int(input("選擇客戶或廠商   :"))
    while(number > len(business_list)):
            number=int(input("選擇客戶或廠商   :"))

    print(business_list[number-1])

    third_temp=[]#日期+名子篩選
    for i in range(len(second_temp)):
        if business_list[number-1] in second_temp[i]:
            third_temp.append(second_temp[i])

    #print(third_temp ) 
    last_total=0
    
    y=3
    last_total=0
    for i in range(len(third_temp)):
        one_list= json.loads(account_temp.get(third_temp[i][0:11]))
        respond=show_form(one_list,i,y,1,last_total)
        last_total=last_total+respond[0]#回傳值total
        y=respond[1]#回傳值y 座標
        last_total=respond[2]#回傳值總計
    print("共計  :",last_total)

if(choose==4):
    def show_goods_form(one_list,flag,goods_temp):
        temp=[]
        goods_temp=goods_temp
        one3_list=one_list[2]
        
        for i in range(len(one3_list)):
            if flag==1:
                temp.append(one3_list[i-1][0])
                temp.append(one3_list[i-1][1])
                temp.append(0)
                
            else:   
                temp.append(one3_list[i-1][0])
                temp.append(0)
                temp.append(one3_list[i-1][1])

            goods_temp.append(temp)
            temp=[]
        return   goods_temp                
                












                
            
            
        

    customer=["協生","滿城","三洋","協生","谷王","尤順緯","何明洲"]

    print("this is 4")
    account_temp=[]
    url='account/'
    account_temp=(readdate(url))#表單資料
    url='inguire/'
    account_name_temp=(readdate(url))#表單資料+name
    

    title=[]
    for  hh in account_name_temp:
        title.append(hh) #表單標籤資料有名字      

    start=input("查詢資料 開始日期 ex 20200413 :  ")
    while start.isdigit()!=True or isValidDate(start)  !=True:
          start=input("查詢資料 開始日期 ex 20200413 :  ")
    end=input("查詢資料 (結束日期) ex 20200413 :  ")
    while end.isdigit()!=True  or isValidDate(end)  !=True:
          end=input("查詢資料 (結束日期) ex 20200413 :  ")
   
    day_choose=geteveryday(start,end)#取得所有日子
    
    second_temp=[]
    for i in range(len(day_choose)):
        for j in range(len(title)):
            if(day_choose[i] in title[j]):
                second_temp.append(title[j])#日期篩選

    goods_temp=[]
    for i in range(len(second_temp)):
        one_list= json.loads(account_temp.get(second_temp[i][0:11]))
        correct=1
    
        for  j in range(len(customer)):
            if second_temp[i][12:13] in customer[j]:#廠商或顧客
                 correct=2                  
                 break
        goods_temp=show_goods_form(one_list,correct,goods_temp)

   
    goods_temp.sort(key=lambda goods_temp:goods_temp[0])#陣列排序
    print()
    

    goods_last_temp=[]
    for i in range(len(goods_temp)):
        if i==0:
            index=i#陣列所在位址
            
            
            goods_last_temp.append(goods_temp[i])
            head=goods_last_temp[0][0]#新陣列 第一品名
            
            #print("head" ,goods_last_temp,head)
        else:
            if head==goods_temp[i][0]:
                goods_last_temp[index][1]=float(goods_last_temp[index][1])+float(goods_temp[i][1])
                goods_last_temp[index][2]=float(goods_last_temp[index][2])+float(goods_temp[i][2])
            else:
                goods_last_temp.append(goods_temp[i])
                index=index+1
                head=goods_last_temp[index][0]#新陣列 第一品名
                
    #print(len(goods_last_temp) )          
    import os
    os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

    import openpyxl
    wb = openpyxl.load_workbook('goods_use.xlsx', data_only=True)

    s1 = wb['sheet1']            # 開啟工作表 1

    s1.cell(1,1).value = '產   品   名   稱'# 儲存格 A1 內容為 apple
    s1.cell(1,2).value = '進 貨'
    s1.cell(1,3).value = '出 貨'
    
    y=2
    for i in range(len(goods_last_temp)):
        print(goods_last_temp[i])
        s1.cell(y,1).value =goods_last_temp[i][0]
        s1.cell(y,2).value =goods_last_temp[i][1]
        s1.cell(y,3).value =goods_last_temp[i][2]
        y=y+1
    wb.save('goods_use.xlsx')

        
