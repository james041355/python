# 引用必要套件
import firebase_admin
from firebase_admin import credentials
#from firebase_admin import firestore
from firebase_admin import db
#import matplotlib.pyplot as plt
#import pandas as pd
#import numpy as np
import json
import datetime
from datetime import date
import tkinter as tk
# 引入訊息視窗模組
import tkinter.messagebox

# 引用私密金鑰
# path/to/serviceAccount.json 請用自己存放的路徑
cred = credentials.Certificate('/Users/home/Desktop/python/outdata-6457c-firebase-adminsdk-2miud-7b496d14a7 (1).json')

# 初始化firebase，注意不能重複初始化
firebase_admin.initialize_app(cred,{
	'databaseURL':'https://outdata-6457c.firebaseio.com'})


def readdate(dateset):#讀取(位置+標籤)####
    ref=db.reference(dateset)
    #print(ref.get())                 
    return(ref.get())

def update(dateset,datenames,date):##建立 位置/標籤/資料/
    ref=db.reference(dateset)
    ref.update({datenames:date})

def deldate(dateset):#刪除位置(位置+標籤)
    ref=db.reference(dateset)
    ref.set({})

def show_goods(Atemp):#產品顯示####
    print()
    for i in range(len(Atemp)):
        if i==1:
            print(list1[i],"-----",Atemp[i]
                ,"______","加稅價","-----","{:.2f}".format(float(Atemp[i])*1.05))
        else:
            print(list1[i],"-----",Atemp[i])


def firebase_get(a):#拿firebase資料####
    date_temp = json.loads(readdate(a))
    #print(date_temp)
    return(date_temp)

def show_number(a,fix):#尋找標號 及 標籤###
    a=a
    fix=fix
    for i in range(2,goods_a+1):
        d1=date_temp.get(str(i))
        d2=json.loads(d1)
        if(a==d2[4]):
            lable=str(i)#lable標籤
            one_goods_temp=d2#資料陣列 是本機資料 非電腦資料
            #print(one_goods_temp)
            #print(d2,fix,path_goods,lable)
            show_one_good(d2,fix,path_goods,lable)
            #show_goods(d2) 
    
def show_one_good(Atemp,fix,path,lable):#印出品項與修改### 
    path=path
    lable=lable
    fix=fix
    Atemp=Atemp
    #print(Atemp,fix,path_goods,lable)    
    if fix==0:
        for i in range(len(Atemp)):
            if i==1:
                print(list1[i],"-----",Atemp[i],"______",
                    "加稅價","-----","{:.2f}".format(float(Atemp[i])*1.05))
            else:
                print(list1[i],"-----",Atemp[i])
    else:
        for i in range(len(Atemp)):
            if i==1:
                print(list1[i],"-----",Atemp[i],"______",
                    "加稅價","-----","{:.2f}".format(float(Atemp[i])*1.05))
            else:
                print(list1[i],"-----",Atemp[i])
        print("\n需要修改  1: 要修改 2 :不須修改  3 :刪除產品資料   ",end="")
        
        number_b=(input())
        #print("keyin=",type(number_b))
        while number_b.isdigit()!=True :
            print("\n需要修改  1: 要修改 2 :不須修改  3 :刪除產品資料   ",end='')
            number_b=(input())
        number_b=int(number_b)
        if number_b==1   :
            for i in range(len(Atemp)):
                if i==1:
                    print(i+1,":",list1[i],"-----",Atemp[i],"______",
                        "加稅價","-----","{:.2f}".format(float(Atemp[i])*1.05))
                else:
                    print(i+1,":",list1[i],"-----",Atemp[i])
            
            print("\n需要修改的選項  選 9 : 結束   ",end="")
            
            number=(input())
            while number.isdigit()!=True:
                print("\n需要修改的選項  選 9 : 結束   ",end='')
                number=input()
            number=int(number)
            if number==9:

                show_one_good(Atemp,0,path,lable)
            else:    
                print(list1[number-1],"-----","(",Atemp[number-1],")  ",end="")
                Atemp[number-1]=input()
                print()
                show_one_good(Atemp,1,path,lable)
        elif number_b==3:
        
            print("刪除資料  1: 確定刪除  2: 放棄刪除    ",end="")  
            number_a=(input())
            while number_a.isdigit()!=True:
                print("刪除資料  1: 確定刪除  2: 放棄刪除    ",end='')
                number_a=input()
            #print(path_goods,lable,goods_a)
            #print (path_goods,type(lable),type(goods_a))
            number_a=int(number_a)
            if number_a==1:
                
                if int(lable)==int(goods_a):
                    deldate(path_goods+lable)
                    update(path_goods,'a',goods_a-1)
                    print("del_date")
                else:
                    
                    temp=readdate(path_goods+str(goods_a))
                    #print(temp)
                    
                    update(path_goods,lable,temp)
                    deldate(path_goods+str(goods_a))
                    update(path_goods,'a',goods_a-1)
                    print("del_date")
        else:                        
            show_one_good(Atemp,0,path,lable)
            
            Atemp=json.dumps(Atemp)
            update(path,lable,Atemp)
            print("update")


def geteveryday(begin_date,end_date):#計算日期天數###
    date_list = []
    begin_date = datetime.datetime.strptime(begin_date, "%Y%m%d")
    end_date = datetime.datetime.strptime(end_date,"%Y%m%d")
    while begin_date <= end_date:
        date_str = begin_date.strftime("%Y%m%d")
        date_list.append(date_str)
        begin_date += datetime.timedelta(days=1)
    return date_list

def isValidDate(datestr):#是否日期格式###
    try:
        date.fromisoformat(datestr)
    except:
        return False
    else:
        return True

def show_form(one_list,count,y,x,last_total):#印 交易明細用
        y=y
        x=x
        last_total=last_total
        one3_list=one_list[2]
        print("日期  :",(one_list[0]))
    
        s1.cell(y,x).value =one_list[0]
        x=2      
        if count==0:
            print("客戶  :",(one_list[1]))
            s1.cell(1,1).value ="客戶名稱  " 
            s1.cell(1,2).value =one_list[1]
            s1.cell(2,1).value ="日　期  "            
            s1.cell(2,2).value ="產　品　名　稱"           
            s1.cell(2,3).value ="數量"          
            s1.cell(2,4).value ="單位"         
            s1.cell(2,5).value ="單價"          
            s1.cell(2,6).value ="小計"

        total=0
        for i in range(len(one3_list)):
            total=total+one3_list[i-1][4]
            print("品名:",'{:15}'.format(one3_list[i-1][0]) ,'%-5s' %one3_list[i-1][1],one3_list[i-1][2],
            '單價','%-5s' %one3_list[i-1][3],"小計:",'%-5s' %one3_list[i-1][4])
            
            s1.cell(y,x).value =one3_list[i-1][0]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][1]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][2]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][3]
            x=x+1
            s1.cell(y,x).value =one3_list[i-1][4]
            y=y+1
            x=2
        last_total=last_total+total
        s1.cell(y,1).value="總計"
        s1.cell(y,6).value=last_total

        print ("總計: ",total)
        print("-------------------------------------------------------")
        wb.save('form_use.xlsx')
        return total,y,last_total


    
def show_goods_form(one_list,flag,goods_temp):#計算進出貨
        temp=[]
        goods_temp=goods_temp
        one3_list=one_list[2]
        
        for i in range(len(one3_list)):
            if flag==1:
                temp.append(one3_list[i-1][0])
                temp.append(one3_list[i-1][1])
                temp.append(0)
                
            else:   
                temp.append(one3_list[i-1][0])
                temp.append(0)
                temp.append(one3_list[i-1][1])

            goods_temp.append(temp)
            temp=[]
        return   goods_temp

def show_one_goods(one_list,name,goods_temp):#找單一產品在表單的數量####
        temp=[]
        name=name
        goods_temp=goods_temp
        
        one2=one_list[1]
        one3_list=one_list[2]
        
        for i in range(len(one3_list)):
            
            if name== one3_list[i][0] :
                temp.append(one_list[0])
                temp.append(one_list[1])
                temp.append(one3_list[i][1])
                temp.append(one3_list[i][2])
                temp.append(one3_list[i][3])
                goods_temp.append(temp)
                temp=[]
        return   goods_temp

def check_goods(number,flag):#有相同產品名稱或相同編號####
    number=number
    flag=flag
    for i in range(2,goods_a+1):
                d1=date_temp.get(str(i))
                d2=json.loads(d1)
                if(number == d2[0] or number == d2[4] ):
                    print("有相同產品名稱或相同編號")
                    #print(list1[0],"  :",end='')
                    #number=input()
                    flag =1
                    print(i)
                    break
    return flag
    
def wire_show_name(name,cost,sell):#尋找成本 及 標籤####
            name=name
            cost=cost
            sell=sell
            for i in range(2,goods_a+1):
                d1=date_temp.get(str(i))
                d2=json.loads(d1)
                if(name==d2[0]):
                    lable=str(i)#lable標籤
                    d2[1]=int(cost) #本錢
                    d2[2]=int(sell)#大賣
                    d2[3]=int(sell)#小賣
                    d2=json.dumps(d2)
                    update(path_goods,lable,d2)
                    #print(path_goods,lable,d2)
                    print("update")

def have_Date(name):
    try:
        json.loads(date_temp.get(name))
    except:
        return False
    else:
        return True
def show_form_window(date,serial,guest,temp):#日期,序號,廠商名,資料
    date=date
    serial=serial
    guest=guest
    temp=temp
    #print(len(temp))
    window = tk.Tk()

    window.title('輸入表單')#程式上方的文字
    window.geometry("600x600+750+50")
    ## 設定視窗大小為 300x100，視窗（左上角）在螢幕上的座標位置為 (250, 150)

    window.resizable(False, False)#定義可不可以被使用者放大縮小視窗
    #window.iconbitmap('icon.ico')# 設定程式的圖示，可在括弧中放入檔案路徑

    

    def onOK():
        window.destroy()#關閉螢幕
    def onnext():
        keep_temp.append(1)
        window.destroy()#關閉螢幕    
    def onfix():
        keep_temp.append(2)
        window.destroy()#關閉螢幕
    def onsave():
        keep_temp.append(3)
        window.destroy()#關閉螢幕
    def onfixsave():
        keep_temp.append(4)
        window.destroy()#關閉螢幕       

    # 以預設方式排版標示文字

    label_date=tk.Label(window, text = '日期  ',font = ('Arial', 14))
    label_date_in=tk.Label(window, text = date,font = ('Arial', 14))
    label_serial=tk.Label(window, text = '序號  ',font = ('Arial', 14))
    label_serial_in=tk.Label(window, text = serial,font = ('Arial', 14))
    label_name=tk.Label(window, text = '店家名字  ',font = ('Arial', 14))
    label_name_in=tk.Label(window, text = guest,font = ('Arial', 14))
    x=0
    y=0
    label_date.place(x=x,y=y)
    x=x+50
    label_date_in.place(x=x,y=y)
    x=x+100
    label_serial.place(x=x,y=y)
    x=x+50
    label_serial_in.place(x=x,y=y)
    x=x+150
    label_name.place(x=x,y=y)
    x=x+100
    label_name_in.place(x=x,y=y)

    # 標示文字
    labela = tk.Label(window, text = '產   品   名   稱  ',font = ('Arial', 14))
    labelb = tk.Label(window, text = '數 量  ',font = ('Arial', 14))
    labelc = tk.Label(window, text = '單 位  ',font = ('Arial', 14))
    labeld = tk.Label(window, text = '單 價  ',font = ('Arial', 14))
    labele = tk.Label(window, text = '小 計  ',font = ('Arial', 14))
    labelf = tk.Label(window, text = '備 註  ',font = ('Arial', 14))
    #labela.pack(side='left')
    x=0
    y=50
    labela.place(x=x,y=y)
    x=x+200
    labelb.place(x=x,y=y)
    x=x+70
    labelc.place(x=x,y=y)
    x=x+70
    labeld.place(x=x,y=y)
    x=x+70
    labele.place(x=x,y=y)
    x=x+70
    labelf.place(x=x,y=y)
    x=0
    y=100
    total=0
    for i in range(len(temp)):
        label = tk.Label(window, text = temp[i][0],font = ('Arial', 14))
        label.place(x=x,y=y)
        x=x+200
        label = tk.Label(window, text = temp[i][1],font = ('Arial', 14))
        label.place(x=x,y=y)
        x=x+70
        label = tk.Label(window, text = temp[i][2],font = ('Arial', 14))
        label.place(x=x,y=y)
        x=x+70
        label = tk.Label(window, text = temp[i][3],font = ('Arial', 14))
        label.place(x=x,y=y)
        x=x+70
        label = tk.Label(window, text = temp[i][4],font = ('Arial', 14))
        label.place(x=x,y=y)
        total=total+temp[i][4]
        #total=0
        x=0
        y=y+25

    x=0
    y=y+25
    label = tk.Label(window, text = "----------------------------------------------------------------------------------------------------------")
    label.place(x=x,y=y)
    x=480
    y=y+25   
    label = tk.Label(window, text = total,font = ('Arial', 14))
    label.place(x=x,y=y)
    
    button_next = tk.Button(window, text = "下一筆資料", command = onnext,bg='gray')
    button_next.pack(side='bottom',anchor=tk.SW)
    button_fix = tk.Button(window, text = "修改單筆資料", command = onfix,bg='gray')
    button_fix.pack(side='bottom',pady=5,anchor=tk.SW,)
    button_save = tk.Button(window, text = "儲存表單", command = onsave,bg='yellow')
    button_save.pack(side='bottom',pady=5,anchor=tk.SW)
    button_fix_save = tk.Button(window, text = "修改成本", command = onfixsave,bg='gray')
    button_fix_save.pack(side='bottom',pady=5,anchor=tk.SW)

    window.mainloop()

def show_to_window(temp,choose):#來源資料,(1 客戶名稱 2 產品用),
            temp=temp
            choose=choose
            window = tk.Tk()
            if choose==1:
                window.title('選擇 廠商/客戶 名稱')#程式上方的文字
                window.geometry("400x700+250+150")
            else:
                window.title('選擇 產品 名稱')#程式上方的文字
                window.geometry("400x700+250+50")
            

            def onOK():
                keyin=entry.get()
                count=(len(temp))
                #print(keyin[0])
                if keyin != ""  and keyin.isdigit()==True:
                    keyin=int(keyin)
                
                    #print(keyin,type(keyin),"    ",count,type(count))
                    if keyin > count  :
                        
                        if keyin ==99 :
                            keep_temp.append(keyin)
                            window.destroy()#關閉螢幕
                        else:#print("wrong")
                            tkinter.messagebox.showinfo(title = '選擇 廠商/客戶 名稱', # 視窗標題
                                        message = "wrong ")   # 訊息內容
                    else:
                        #print("Hello, {}.".format(entry.get()))
                        if choose==1:
                            print("客戶名稱 :",temp[keyin-1])
                        else:   
                            print("名稱或編號 :",temp[keyin-1][0])

                        keep_temp.append(temp[keyin-1])
                        window.destroy()#關閉螢幕
                else:
                    tkinter.messagebox.showinfo(title = '選擇 廠商/客戶 名稱', # 視窗標題
                                        message = "wrong ")   # 訊息內容

            
            window.wm_attributes('-topmost',1)
            
            if choose == 1:
                for i in range(len(temp)):
                    index=str(i+1)
                    date=str(temp[i])
                    
                    label = tk.Label(window, text = index+date,font = ('Arial', 14))
                    label.pack()
            else:
                for i in range(len(temp)):
                    index=str(i+1)+"  :"
                    date=str(temp[i][0]) +"--"+str(temp[i][4])
                    
                    label = tk.Label(window, text = index+date,font = ('Arial', 14))
                    label.pack()
            
                label = tk.Label(window, text = ' 99   :重新選擇 ',font = ('Arial', 14))
                label.pack()

            entry = tk.Entry(window, width = 20,bd=5,font = ('Arial', 18))    # 輸入欄位所在視窗  # 輸入欄位的寬度                    
            entry.pack()
            button = tk.Button(window, text = "OK", command = onOK)
            button.pack()
            window.mainloop()

def check_enter(name):#防 enter 鍵
    name=name
    if name != "":
        return int(name)

#程式開始的地方
choose=0
while choose != 88:

    keep_temp=[]
    date_temp=[]
    path_goods='tbuiness/'#產品與電線目錄
    date_temp=(readdate(path_goods))#產品與電線用
    business_b= int(date_temp.get('b'))#客戶數量
    factory_f= int(date_temp.get('f'))#廠商數量
    goods_a= int(date_temp.get('a'))#產品數量

    ####拿到所有廠商客戶名子
    business_list=[]
    for i in range(1001,int(factory_f)+1):
        #print(json.loads(date_temp.get(str(i)))[1])
        business_list.append(json.loads(date_temp.get(str(i)))[1])
    for i in range(1101,int(business_b)+1):
        #print(json.loads(date_temp.get(str(i)))[1])
        business_list.append(json.loads(date_temp.get(str(i)))[1])
    #########放在business_list
    #print(business_list)
    customer=[]#拿到所有 客戶名子
    for i in range(1101,int(business_b)+1):
        #print(json.loads(date_temp.get(str(i)))[1])
        customer.append(json.loads(date_temp.get(str(i)))[1])

    print()
    print("    主選項")
    print(" 1 :產品查詢  2 :電線查詢  3 :表單查詢  4 :貨品進出  5 :客戶/廠商 設定  88 :結束程式\n")
    choose=(input())
    
    choose=check_enter(choose)
        
    if(choose==1):

        choose_a=0#total=0
        while choose_a != 88:
            list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
            date_temp=[]
            date_temp=(readdate(path_goods))#產品與電線用
            business_b= int(date_temp.get('b'))##客戶數量 1101~~
            factory_f= int(date_temp.get('f'))#廠商數量 1001~~
            goods_a= int(date_temp.get('a'))#產品數量
            print("    產品查詢")
            print(" 1 :產品查詢 ","2 :修改產品資料 ","3 :建立產品資料 ","4 :刪除產品資料 ","88:結束產品查詢\n")
            choose_a=(input()) 
            choose_a=check_enter(choose_a)

            if choose_a==1:
                fix=0
            elif  choose_a==2:
                fix=1
            elif  choose_a==3:
                fix=3
            elif  choose_a==4:
                fix=4
            else:
                fix=5
                

            if fix==0 or fix==1 or fix==4:

                check=0
                while check != 1:
                    number=input("產品查詢名稱或編號 :   ")
                    if number=="":
                        check=0
                    else:
                        check=1




                list_inquire=[] #收集找到的值
                for i in range(2,goods_a+1):
                    d1=date_temp.get(str(i))
                    d2=json.loads(d1)
                    if(number in d2[0] or number in d2[4] ):
                        list_inquire.append(d2)    
                        
                        

                #print(list_inquire)
                if(len(list_inquire)==0):
                    print("找不到資料")   

                elif(len(list_inquire)==1):
                    #print(list_inquire[0][4]) 
                    #print(number) 
                    show_number(list_inquire[0][4],fix)
                    
                else:
                    for i in range(len(list_inquire)):
                        print(i+1,"  ",list_inquire[i-1][0],"(",list_inquire[i-1][4],")")

                
                    number=(input("選項號碼   :"))
                    while number.isdigit()!=True  :
                        print("選項號碼   :",end='')
                        number=input()                    
                    number=int(number)
                    
                    while number > len(list_inquire)  :
                        print("選項號碼   :",end='')
                        number=input()                 
                        number=int(number)
                    
                    show_number(list_inquire[number-2][4],fix)    

                    print('end')
                
            elif fix==3:#建立產品資料
                #print(goods_a)
                one_goods=[]#所有的輸入欄位
                flag=1

                while flag == 1:

                    print(list1[0],"  :",end='')#list1 是設定的欄位定義
                    number=input()
                    flag=check_goods(number,0)
                    #print(flag)
                one_goods.append(number)
                

                print(list1[1],"僅限輸入數字  :",end='')
                number=input()
                while number=="":
                    print(list1[1],"僅限輸入數字  :",end='')
                    number=input()
                number=float(number)
                one_goods.append(number)
                
                print(list1[2],"僅限輸入數字  :",end='')
                number=input()
                while number=="":
                    print(list1[2],"僅限輸入數字  :",end='')
                    number=input()
                number=float(number)
                one_goods.append(number)
                
                print(list1[3],"僅限輸入數字  :",end='')
                number=input()
                while number=="":
                    print(list1[3],"僅限輸入數字  :",end='')
                    number=input()
                number=float(number)
                one_goods.append(number)
                
                flag=1

                while flag == 1:

                    print(list1[4],"  :",end='')
                    number=input()
                    flag=check_goods(number,0)
                    #print(flag)
                one_goods.append(number)
                
                print(list1[5],"  :",end='')
                number=input()
                one_goods.append(number)
                while number.isdigit()!=True:
                    print(list1[6],"僅限輸入數字  :",end='')
                    number=input()
                number=int(number)
                one_goods.append(number)
                one_goods.append(one_goods[0])
                show_goods(one_goods)
                print(one_goods)
                

                print("確定要存檔嗎  1: 是  2: 不用存檔  ",end='')
                number=(input())
                number=check_enter(number)
                if number==1:
                    one_goods=json.dumps(one_goods)
                    goods_a=goods_a+1
                    #print(goods_a)
                    update(path_goods,goods_a,one_goods)##建立 位置/檔名/資料/
                    update(path_goods,'a',goods_a)
                    print("update")
                else:
                    print("沒存檔")
            else:
                choose_a=88


    if(choose==2):
        print("請先關閉  wire_use.xlsx")
        number=0
        while number != 88:
            path_goods='tbuiness/'#產品與電線目錄
            date_temp=(readdate(path_goods))#產品與電線用
            business_b= int(date_temp.get('b'))#客戶數量
            factory_f= int(date_temp.get('f'))#廠商數量
            goods_a= int(date_temp.get('a'))#產品數量
            
            test=(have_Date('wire'))
            if test==False:
                empty=0
                print("    無電線設定檔")
            else:
                wire_a= json.loads(date_temp.get('wire'))#產品數量
                empty=1
            
            
            wire=[["2平方二芯電纜線",1089,1],
            ['2平方三芯電纜線',1425,1],
            ['3.5平方二芯電纜線',1520,1],
            ['3.5平方三芯電纜線',2079,2],
            ['5.5平方二芯電纜線',2235,2],
            ['5.5平方三芯電纜線',3076,2],
            ['8平方二芯電纜線',3106,2],
            ['8平方三芯電纜線',4313,2],
            ['14平方三芯電纜線',7436,3],
            ['1.6二芯白扁線',700,3],
            ['2.0二芯白扁線',1000,3],
            ['1.6單芯電線',232,4],
            ['2.0單芯電線',350,4],
            ['3.5平方電線',459,4],
            ['5.5平方電線',708,4],
            ['8平方電線',1009,4]
            ]
            print("      電線查詢")
            print(" 1: 查詢電線價錢與列印  2: 電線設定檔 3: 更改電線成本 88: 結束電線查詢\n  ")
            number=(input())
            number=check_enter(number)

            if number==1 and empty != 0:
                for i in range(len(wire_a)):
                    print(i+1,"  ",wire_a[i])
                number_a=(input("選擇月份項   :"))
                while number_a.isdigit()!=True  :
                        print("選擇月份項   :",end='')
                        number_a=input()
                number_a=int(number_a)
                while number_a > len(wire_a)  :
                        print("選擇月份項   :",end='')
                        number_a=input()
                        number_a=int(number_a)
                
                temp=wire_a[number_a-1]
                # title = ["名稱          ","成本","大賣"]
                #print(title)
                import os
                os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

                import openpyxl
                wb = openpyxl.load_workbook('wire_use.xlsx', data_only=True)

                s1 = wb['sheet1']            # 開啟工作表 1
                s1['A1'].value = '名稱'     # 儲存格 A1 內容為 apple
                s1['B1'].value = '成本'    # 儲存格 A2 內容為 orange
                s1['C1'].value = '大賣'    # 儲存格 A3 內容為 banana

                for i in range(len(wire)):
                    if(wire[i][2]==1):
                        para=int(temp[1])
                    elif(wire[i][2]==2):
                        para=int(temp[2])
                    elif(wire[i][2]==3):
                        para=int(temp[3])
                    else:
                        para=int(temp[4])    

                    print("名稱  :",wire[i][0],
                        "成本  :",int(int(wire[i][1])*para/100),
                        "售價  :",int(int(wire[i][1])*(para+10)/100)
                            )
                    print()            
                    s1.cell(i+2,1).value = wire[i][0]      # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
                    s1.cell(i+2,2).value = int(int(wire[i][1])*para/100)    # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
                    s1.cell(i+2,3).value = int(int(wire[i][1])*(para+10)/100)     # 儲格 B3 內容 ( row=3, column=2 ) 為 300
                
                
                wb.save('wire_use.xlsx')
            
            if number==2:
                print("\n 1: 建立電線設定檔  2: 刪除電線設定檔  ")
                choose_a=(input())
                while choose_a.isdigit()!=True  :
                        print("\n 1: 建立電線設定檔  2: 刪除電線設定檔  ",end='')
                        choose_a=input()
                choose_a=int(choose_a)
                if choose_a==1:
                    temp=[]
                    test2=[]

                    print("\n設定年月 ex (11201) ",end="")
                    keyin=input()
                    temp.append(keyin)
                    
                    print("\n 3.5 二芯以下設定 ex (212) ",end="")
                    keyin=input()
                    while keyin.isdigit() != True:
                        print("\n 3.5 二芯以下設定 ex (212) ",end="")
                        keyin=input()
                    temp.append(keyin)
                    print("\n 8平方 四芯以下設定 ex (212) ",end="")
                    keyin=input()
                    while keyin.isdigit() != True:
                        print("\n 8平方 四芯以下設定 ex (212) ",end="")
                        keyin=input()
                    temp.append(keyin)
                    print("\n 白扁線設定 ex (212) ",end="")
                    keyin=input()
                    while keyin.isdigit() != True:
                        print("\n 白扁線設定 ex (212) ",end="")
                        keyin=input()
                    temp.append(keyin)
                    print("\n 22平方電線設定 ex (212) ",end="")
                    keyin=input()
                    while keyin.isdigit() != True:
                        print("\n 22平方電線設定 ex (212) ",end="")
                        keyin=input()
                    temp.append(keyin)
                    #temp=json.dumps(temp)
                    if empty != 0:
                        wire_a.append(temp)
                        wire_a=json.dumps(wire_a)
                        #print(wire_a)
                        update(path_goods,'wire',wire_a)
                    else:
                        test2.append(temp)
                        test2=json.dumps(test2)
                        update(path_goods,'wire',test2)
                        print("update")
                else:
                    if empty != 0:
                        for i in range(len(wire_a)):
                            print(i+1,"   ",wire_a[i])
                        print("\n 選擇第幾項   ",end='')
                        choose_a=(input())
                        while choose_a.isdigit()!=True  :
                            print("\n 選擇第幾項   ",end='')
                            choose_a=input()
                        choose_a=int(choose_a)
                        
                        
                        del wire_a[choose_a-1]

                        #print ("less",len(wire_a))
                        if len(wire_a)==0:
                            deldate(path_goods+'wire')
                        else:
                            wire_a=json.dumps(wire_a)
                            update(path_goods,'wire',wire_a)
            if number==3: 
            
                for i in range(len(wire_a)):
                    print(i+1,"  ",wire_a[i])
                number_a=(input("選擇月份項   :"))
                while number_a.isdigit()!=True  :
                    print("選擇月份項   :",end='')
                    number_a=input()
                number_a=int(number_a)
                while number_a > len(wire_a)  :
                    print("選擇月份項   :",end='')
                    number_a=input()
                    number_a=int(number_a)
                
                temp=wire_a[number_a-1]
                print(temp)
                
                for i in range(len(wire)):
                    if(wire[i][2]==1):
                        price=int(temp[1])*wire[i][1]/100
                        sell=(int(temp[1])+10)*wire[i][1]/100
                        print(wire[i][0],"本錢是  ",price,"小賣 "  ,sell)
                        wire_show_name(wire[i][0],price,sell)
                    elif(wire[i][2]==2):
                        price=int(temp[2])*wire[i][1]/100
                        sell=(int(temp[2])+10)*wire[i][1]/100
                        print(wire[i][0],"本錢是  ",price,"小賣 "  ,sell)
                        wire_show_name(wire[i][0],price,sell)
                    elif(wire[i][2]==3):
                        price=int(temp[3])*wire[i][1]/100
                        sell=(int(temp[3])+10)*wire[i][1]/100
                        print(wire[i][0],"本錢是  ",price,"小賣 "  ,sell)
                        wire_show_name(wire[i][0],price,sell)
                    else:
                        price=int(temp[4])*wire[i][1]/100
                        sell=(int(temp[4])+10)*wire[i][1]/100
                        print(wire[i][0],"本錢是  ",price,"小賣 "  ,sell)
                        wire_show_name(wire[i][0],price,sell)
            else:
                number=88

    if(choose==3):
        def new_creat_goods():#建立新產品 輸入 james 觸發
            list1 = ["名稱","成本","大賣","小賣","編號","單位","數量","簡稱"]
            goods_a= int(date_temp.get('a'))#產品數量
            print("1 : 建立新產品 \n")

                #print(goods_a)
            one_goods=[]
            flag=1
            while flag == 1:
                print(list1[0],"  :",end='')
                number=input()
                flag=check_goods(number,0)
                    #print(flag)
            one_goods.append(number)
                    
            print(list1[1],"僅限輸入數字  :",end='')
            number=input()
            while number.isdigit()!=True:
                print(list1[1],"僅限輸入數字  :",end='')
                number=input()
            number=float(number)
            one_goods.append(number)
                    
            print(list1[2],"僅限輸入數字  :",end='')
            number=input()
            while number.isdigit()!=True:
                print(list1[2],"僅限輸入數字  :",end='')
                number=input()
            number=float(number)
            one_goods.append(number)
                    
            print(list1[3],"僅限輸入數字  :",end='')
            number=input()
            while number.isdigit()!=True:
                print(list1[3],"僅限輸入數字  :",end='')
                number=input()
            number=float(number)
            one_goods.append(number)
                    
                
            flag=1

            while flag == 1:

                print(list1[4],"  :",end='')
                number=input()
                flag=check_goods(number,0)
                    #print(flag)
            one_goods.append(number)
                    
            print(list1[5],"  :",end='')
            number=input()
            one_goods.append(number)
            while number.isdigit()!=True:
                print(list1[6],"僅限輸入數字  :",end='')
                number=input()
            number=int(number)
            one_goods.append(number)
            one_goods.append(one_goods[0])
                
            print(one_goods)

            print("確定要存檔嗎  1: 是  2: 不用存檔  ",end='')
            number=(input())
            while number.isdigit()!=True:
                print("確定要存檔嗎  1: 是  2: 不用存檔  ",end='')
                number=input()
            number=int(number)
            if number==1:
                one_goods=json.dumps(one_goods)
                goods_a=goods_a+1
                    #print(path_goods)
                    #print(goods_a)
                    #print(one_goods)
                    
                update(path_goods,goods_a,one_goods)##建立 位置/檔名/資料/
                update(path_goods,'a',goods_a)
                print("update")
                    
            else:
                print("沒存檔")
                
                
                
            print(" 產品建立完畢  繼續表單輸入")
            number=input("名稱或編號 :   ")
            return number,goods_a
#############副程式############

        ### 第三選項開始

        choose_a=0
        while choose_a != 88:
            print("一定要關閉  form_use.xlsx")
            import os
            os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

            import openpyxl
            wb = openpyxl.load_workbook('form_use.xlsx', data_only=True)

            s1 = wb['sheet1']            # 開啟工作表 1

            account_temp=[]
            path_form='account/'
            account_temp=(readdate(path_form))#表單資料
            path_form_name='inguire/'
            account_name_temp=(readdate(path_form_name))#表單資料+name
            

            title=[]
            for  hh in account_name_temp:
                title.append(hh) #表單標籤資料有名字 
            print("      表單查詢")
            print(" 1: 表單查詢(不可修改)   2: 增加/刪除 表單   88:結束表單作業   \n")
            choose_a=(input())
            choose_a=check_enter(choose_a)
            if choose_a==1:
                start=input("查詢資料 開始日期 ex 20200413 :  ")
                while isValidDate(start)  !=True:
                    start=input("查詢資料 開始日期 ex 20200413 :  ")
                end=input("查詢資料 (結束日期) ex 20200413 :  ")
                while isValidDate(end)  !=True:
                    end=input("查詢資料 (結束日期) ex 20200413 :  ")

                day_choose=geteveryday(start,end)#取得所有日子
                
                second_temp=[]
                for i in range(len(day_choose)):
                    for j in range(len(title)):
                        if(day_choose[i] in title[j]):
                            second_temp.append(title[j])#日期篩選

                #print(second_temp)
                for i in range(len(business_list)):
                    print(i+1,"  ",business_list[i])
                number=(input("選擇客戶或廠商   :"))
                while number.isdigit()!=True  :
                    print("選擇客戶或廠商   :",end='')
                    number=input()
                number=int(number)
                while number > len(business_list)  :
                    print("選擇客戶或廠商   :",end='')
                    number=input()
                    number=int(number)

                print(business_list[number-1])

                third_temp=[]#日期+名子篩選
                for i in range(len(second_temp)):
                    if business_list[number-1] in second_temp[i]:
                        third_temp.append(second_temp[i])

                #print(third_temp ) 
                last_total=0
                
                y=3
                last_total=0
                for i in range(len(third_temp)):
                    one_list= json.loads(account_temp.get(third_temp[i][0:11]))
                    respond=show_form(one_list,i,y,1,last_total)
                    last_total=last_total+respond[0]#回傳值total
                    y=respond[1]#回傳值y 座標
                    last_total=respond[2]#回傳值總計
                print("共計  :",last_total)
            
            if choose_a==2:
                #show_form_window(date,serial,guest,temp):#日期,序號,廠商名,資料
                #show_to_window(temp,choose):#來源資料,(1 客戶名稱 2 產品用),
                fix=0
                head=[]
                date_temp=(readdate(path_goods))#產品與電線用``
                business_b= int(date_temp.get('b'))##客戶數量 1101~~
                factory_f= int(date_temp.get('f'))#廠商數量 1001~~
                goods_a= int(date_temp.get('a'))#產品數量
                serial_net=int(date_temp.get('s'))#序號數字
                if serial_net==99:
                    serial_net=11
                    update(path_goods,'s',serial_net)
                
                #print("序號",serial_net)
                #print(customer)
                
                
                print("\n 1: 增加新表單  2: 單日表單(修改/刪除)查詢 \n    ",end='')
                keyin=(input())
                while keyin.isdigit()!=True:
                        print("\n 1: 增加新表單  2: 單日表單(修改/刪除)查詢\n    ",end='')
                        keyin=(input())
                keyin=int(keyin)
                if keyin==1:
                    creat=1
                    print("   增加新表單:     ")
                else:
                    creat=2
                    print("2: 單日表單(修改/刪除)查詢    ")
                start=input("日 期 ex 20200413 :  ")
                while isValidDate(start)  !=True:
                    start=input("日 期 ex 20200413  :  ")
                if creat==1:
                    head.append(start)#日期
                    serial="N"+str(start)+str(serial_net)#序號
                    show_to_window(business_list,1)#客戶名
                    
                    #print(keep_temp)
                    #kk=input()
                    head.append(keep_temp[0]) #客戶 或 廠商
                    for i in range(len(customer)):
                        if keep_temp[0]==customer[i]:
                            flag=1#客戶
                            break
                        else:
                            flag=2
                    #print(flag)
                    keep_temp=[]
                    #print(head,serial)
                else:
                    day_choose=start
                    day_second_temp=[]
                
                    for j in range(len(title)):
                        if(day_choose in title[j]):
                            day_second_temp.append(title[j])#日期篩選
                    if len(day_second_temp) != 0:###################fix fix fix
                        for j in range(len(day_second_temp)):
                            print(j+1," :",day_second_temp[j])
                        
                        keyin=(input("輸入選擇項  :  "))
                        while keyin.isdigit()!=True :
                            print("輸入選擇項  :  ",end='')
                            keyin=input()
                        keyin=int(keyin)
                        name=day_second_temp[keyin-1]
                        #print("hhh",name[0:11])
                        temp=json.loads(account_temp.get(name[0:11]))   
                        
                        print("\n日期:",temp[0])
                        print("序號:",name[0:11])
                        print("廠商名:",temp[1])
                        for i in range(len(temp[2])):
                            print(temp[2][i])
                        #show_form_window(temp[0],name[0:11],temp[1],temp[2])#日期,序號,廠商名,資料
                        print("\n 1: 修改表單  2: 刪除表單 3: 結束表單    ",end='')
                        keyin_a=(input())
                        
                        while keyin_a.isdigit()!=True:
                            print("\n 1: 修改表單  2: 刪除表單 3: 結束表單    ",end='')
                            keyin_a=(input())
                        keyin_a=int(keyin_a)
                        
                        if keyin_a==2:
                            #print("刪除選單區")
                            #print(path_form_name)
                            #print(name)
                            #print(path_form)
                            #print("serial=",name[0:11])
                            deldate(path_form_name+name)
                            deldate(path_form+name[0:11])
                            print("deldate")
                            next=0
                        elif keyin_a==3:
                            next=0
                        else:
                            head=[]
                            for i in range(len(customer)):
                                if temp[1]==customer[i]:
                                    flag=1#客戶
                                    break
                                else:
                                    flag=2
                            second_temp=temp[2]
                            head.append(day_choose)#日期
                            head.append(temp[1])#名子
                            serial=name[0:11]
                            #print(head)
                            #print("label==",serial)
                            
                            for i in range(len(second_temp)):
                                print(i+1 ,":   ",second_temp[i])
                            
                            print("選擇第幾筆資料  :  ",end='')
                            keyin=input()
                            
                            while keyin.isdigit()!=True  :
                                print("選擇第幾筆資料  :  ",end='')
                                keyin=input()
                            keyin=int(keyin)
                            while keyin > len(second_temp):
                                print("選擇第幾筆資料  :  ",end='')
                                keyin=input()
                                keyin=int(keyin)
                            #fix_temp=second_temp[keyin-1]
                            
                            fix_label=keyin-1#第幾項
                            print(second_temp[keyin-1])
                            print("修改資料單筆資料選項!!!")
                            fix=99
                            next=1
                    else:
                        print("無資料")
                        next=0

                if creat==1:
                    second_temp=[]#儲存值
                    next=1
                while next != 0:
                
                    bottom=[]#輸入值
                    
                    same=0
                    test=goods_a
                    while same != 1:
                        number=input("名稱或編號 :   ")
                        while number=='james':#創造新產品
                            temp=new_creat_goods()
                            goods_a=temp[1]
                            if test != goods_a:
                                date_temp=(readdate(path_goods))#產品與電線用
                            number=temp[0]
                    
                        #print(number)
                        #kk=input()
                        list_inquire=[]
                        for i in range(2,goods_a+1):
                            d1=date_temp.get(str(i))
                            d2=json.loads(d1)
                            if(number in d2[0] or number in d2[4] ):
                                list_inquire.append(d2)    
                                    
                                    

                        #print(list_inquire)
                        if(len(list_inquire)==0):
                    
                            tkinter.messagebox.showerror(title="資訊告知", message="沒有資料")
                            
                        elif(len(list_inquire)==1):
                            print("名稱或編號 :  ",list_inquire[0][0]) 
                            keep_temp.append(list_inquire[0])
                            bottom.append(list_inquire[0][0])
                            new_name=list_inquire[0][0]
                            
                            
                            same=1    
                        elif(len(list_inquire) >25 ):
                            print("請縮小查詢範圍")
                            same=0
                        else:
                            
                            show_to_window(list_inquire,2)
                            #print(keep_temp)
                            
                            if keep_temp[0]==99:
                                same=0
                                keep_temp=[]
                            else:
                                bottom.append(keep_temp[0][0])
                                new_name=keep_temp[0][0]
                                same=1

                    
                    #print(new_name)
                    keyin_a=input("數量       :  ")
                    while keyin_a=="":
                        keyin_a=(input("數量       :  "))
                    bottom.append(float(keyin_a))
                    
                    print("單位       : ",keep_temp[0][5])                
                    bottom.append(keep_temp[0][5])      
                    
                    if flag==2:
                        print("單價","(",keep_temp[0][1], ")" , ":  " ,end="")
                        keyin_b=input()
                        while keyin_b=="":
                            print("單價","(",keep_temp[0][1], ")" , ":  " ,end="")
                            keyin_b=input()
                    else:
                        print("單價","(",keep_temp[0][2], ")" , ":  " ,end="")
                        keyin_b=input()
                        while keyin_b=="":
                            print("單價","(",keep_temp[0][2], ")" , ":  " ,end="")
                            keyin_b=input()


                    bottom.append(float(keyin_b))
                    print("小計       : ",float(keyin_a)*float(keyin_b))
                    bottom.append(float(keyin_a)*float(keyin_b))
                    if fix != 99:
                        second_temp.append(bottom)
                    else:
                        #print(fix_label)
                        
                        second_temp[fix_label]=bottom
                        fix=0
                    #print('serial==',serial)    
                    #print("head==" ,head)
                    #print("bottom==",second_temp)
                    #print("   last end test") 
                    show_form_window(head[0],serial,head[1],second_temp)#日期,序號,廠商名,資料
                    #print("test",keep_temp[1])
                    keyin=keep_temp[1]
                    keep_temp=[]
                    
                    if keyin==3:#結束區
                        #show_form_window(head[0],serial,head[1],second_temp)#日期,序號,廠商名,資料
                        #print('serial==',serial)    
                        
                        #print("bottom==",second_temp)
                        newname=serial+head[1]
                        #print(newname)
                        head.append(second_temp)
                        head=json.dumps(head)
                        #print("head==" ,head)
                        update(path_form,serial,head)
                        update(path_form_name,newname,serial)
                        update(path_goods,'s',serial_net+1)
                        print("update")
                        next=0
                    
                    elif keyin==1:
                        #show_form_window(head[0],serial,head[1],second_temp)#日期,序號,廠商名,資料
                        next=1
                    elif keyin==2:
                        for i in range(len(second_temp)):
                            print(i+1 ,":",second_temp[i])
                        keyin=(input("選擇第幾筆資料    : "))
                        while keyin.isdigit()!=True :
                            print("選擇第幾筆資料  :  ",end='')
                            keyin=(input())
                        keyin=int(keyin)
                        while keyin > len(second_temp)  :
                            print("選擇第幾筆資料  :  ",end='')
                            keyin=(input())
                            keyin=int(keyin)
                        fix_label=keyin-1#第幾項
                        print(second_temp[keyin-1])
                        print("修改資料單筆資料選項!!!")
                        fix=99
                        next=1
                    else:#fix goos
                        print(new_name)
                        list_inquire=[]
                        for i in range(2,goods_a+1):
                            d1=date_temp.get(str(i))
                            d2=json.loads(d1)
                            if(new_name in d2[0]  ):
                                lable=str(i)
                                list_inquire.append(d2)
                        temp=list_inquire[0]
                        print("now",temp)#[1]成本 [2]大賣  [3]小賣
                        print("產品名稱 : ",temp[0])
                        
                        print("修改成本","(",temp[1],")  :",end="")
                        keyin=(input())
                        #print(type(keyin))
                        while keyin == "":
                            print("修改成本","(",temp[1],")  :",end="")
                            keyin=(input())
                        keyin=float(keyin)
                        temp[1]=keyin
                        print("修改大賣 ","(",temp[2],")  :",end="")
                        keyin=(input())
                        while keyin == "":
                            print("修改大賣 ","(",temp[2],")  :",end="")
                            keyin=(input())
                        keyin=float(keyin)
                        temp[2]=keyin
                        
                        print("修改小賣 ","(",temp[3],")  :",end="")
                        keyin=(input())
                        while keyin == "" :
                            print("修改小賣 ","(",temp[3],")  :",end="")
                            keyin=(input())
                        keyin=float(keyin)
                        temp[3]=keyin
                        print("new ",temp)
                        #print(lable)
                        #print(path_goods)
                        print("確定要存檔嗎  1: 是  2: 不用存檔  3: 存檔且結束表單 ",end="")
                        number=(input())
                        number=check_enter(number)
                        if number==1 or  number==3:
                            temp=json.dumps(temp)
                            update(path_goods,lable,temp)##建立 位置/檔名/資料/
                            date_temp=(readdate(path_goods))#產品與電線用
                            print("update")
                            if number==3:
                                newname=serial+head[1]
                                head.append(second_temp)
                                head=json.dumps(head)
                                #print("head==" ,head)
                                update(path_form,serial,head)
                                update(path_form_name,newname,serial)
                                update(path_goods,'s',serial_net+1)
                                print("update")
                                next=0
                        else:
                            print("沒存檔")
            else:
                choose_a=88



    if(choose==4):
        print("\n  4 :全品項/單一貨品 進出查詢\n")
        choose_a=0
        while choose_a !=88:
            
            account_temp=[]
            url='account/'
            account_temp=(readdate(url))#表單資料
            url='inguire/'
            account_name_temp=(readdate(url))#表單資料+name
            
            
            title=[]
            for  hh in account_name_temp:
                title.append(hh) #表單標籤資料有名字      
            print("\n    貨品進出")
            print(" 1 :全品項查詢"," 2 :單一品項查詢","88 :結束 貨品進出查詢\n")
            choose_a=(input())
            choose_a=check_enter(choose_a)
            
            if choose_a==1 or choose_a==2:
                print()
                start=input("查詢資料 開始日期 ex 20200413 :  ")
                while isValidDate(start)  !=True:
                    start=input("查詢資料 開始日期 ex 20200413 :  ")
                end=input("查詢資料 (結束日期) ex 20200413 :  ")
                while isValidDate(end)  !=True:
                    end=input("查詢資料 (結束日期) ex 20200413 :  ")
                print()
                day_choose=geteveryday(start,end)#取得所有日子
                
                second_temp=[]
                for i in range(len(day_choose)):
                    for j in range(len(title)):
                        if(day_choose[i] in title[j]):
                            second_temp.append(title[j])#日期篩選

            
            if choose_a==1:

                goods_temp=[]
                for i in range(len(second_temp)):
                    one_list= json.loads(account_temp.get(second_temp[i][0:11]))
                    correct=1
                
                    for  j in range(len(customer)):
                        if second_temp[i][12:13] in customer[j]:#廠商或顧客
                            correct=2                  
                            break
                    goods_temp=show_goods_form(one_list,correct,goods_temp)
                    
            
                goods_temp.sort(key=lambda goods_temp:goods_temp[0])#陣列排序
                print()
                print("一定要關閉  goods_use.xlsx")
                print("\n此選項有列印去 excel\n")
                goods_last_temp=[]
                for i in range(len(goods_temp)):

                    if i==0:
                        index=i#陣列所在位址
                        
                        
                        goods_last_temp.append(goods_temp[i])
                        head=goods_last_temp[0][0]#新陣列 第一品名
                        
                        #print("head" ,goods_last_temp,head)
                    else:
                        if head==goods_temp[i][0]:
                            goods_last_temp[index][1]=float(goods_last_temp[index][1])+float(goods_temp[i][1])
                            goods_last_temp[index][2]=float(goods_last_temp[index][2])+float(goods_temp[i][2])
                        else:
                            goods_last_temp.append(goods_temp[i])
                            index=index+1
                            head=goods_last_temp[index][0]#新陣列 第一品名
                            
                #print(len(goods_last_temp) )          
                import os
                os.chdir('/Users/home/Desktop')  # Colab 換路徑使用

                import openpyxl
                wb = openpyxl.load_workbook('goods_use.xlsx', data_only=True)

                s1 = wb['sheet1']            # 開啟sheet1

                s1.cell(1,1).value = '產   品   名   稱'# 儲存格 A1 內容為 apple
                s1.cell(1,2).value = '進 貨'
                s1.cell(1,3).value = '出 貨'
                
                y=2
                for i in range(len(goods_last_temp)):
                    print(goods_last_temp[i])
                    s1.cell(y,1).value =goods_last_temp[i][0]
                    s1.cell(y,2).value =goods_last_temp[i][1]
                    s1.cell(y,3).value =goods_last_temp[i][2]
                    y=y+1
                wb.save('goods_use.xlsx')
            if choose_a==2:
                
                goods_a= int(date_temp.get('a'))#產品數量
                correct=0          
                list_inquire=[]
                number=input("產品查詢名稱或編號 :   ")
                for i in range(2,goods_a):
                    if(correct!=1):
                        d1=date_temp.get(str(i))
                        d2=json.loads(d1)
                        if(number in d2[0] or number in d2[4] ):
                        #list_inquire.append(i)
                            list_inquire.append(d2)
                            
                            correct=0
                #print(number,len(list_inquire))
                if(len(list_inquire)==0):
                    print("找不到資料")   
                elif(len(list_inquire)==1):
                    print(number) 
                    for i in range(2,goods_a):
                        d1=date_temp.get(str(i))
                        d2=json.loads(d1)
                        if(number==d2[4]):
                            name=(d2[0])
                
                            
                else:
                    for i in range(len(list_inquire)):
                        print(i+1,"  ",list_inquire[i-1][0],"(",list_inquire[i-1][4],")")

                    number=(input("選項號碼   :"))
                    while number.isdigit()!=True :
                            print("選項號碼   :",end='')
                            number=(input())
                    number=int(number)
                    while number > len(list_inquire) :
                            print("選項號碼   :",end='')
                            number=(input())
                            number=int(number)                    
                    name=(list_inquire[number-2][0])    
                #print(list_inquire)
                goods_temp=[]
                if len(list_inquire) != 0:
                    
                    #goods_temp=[]
                    for i in range(len(second_temp)):
                        one_list= json.loads(account_temp.get(second_temp[i][0:11]))
                        goods_temp=show_one_goods(one_list,name,goods_temp)
                

                if len(goods_temp) != 0:
                    print("\n此選項沒有列印去 excel\n")
                    for i in range(len(goods_temp)):
                        print(goods_temp[i])
                else:
                    print("\n沒有資料")   
            else:   
                choose_a=88


    if(choose==5):
        number=0
        while number != 88:
            
            date_temp=(readdate(path_goods))#產品與電線用
            business_b= int(date_temp.get('b'))##客戶數量 1101~~
            factory_f= int(date_temp.get('f'))#廠商數量 1001~~
            goods_a= int(date_temp.get('a'))#產品數量
            
            
            print("    客戶/廠商 設定")
            print(" 1 :編輯廠商資料","    2 :編輯客戶資料")
            print(" 3 :增加廠商資料","    4 :增加客戶資料")
            print(" 5 :刪除廠商資料","    6 :刪除客戶資料")
            print("88 :結束客戶/廠商 設定")
            number=(input())
            number=check_enter(number)
            if number==1 or number==2 or  number==5 or  number==6:
                for i in range(1001,int(factory_f)+1):
                    temp=(json.loads(date_temp.get(str(i))))
                    print(i-1000,"編號:",temp[0],"名稱:",temp[1],"TEL :",temp[2])
                    keep=i-1000
                
                #print(keep)
                for i in range(1101,int(business_b)+1):
                    temp=(json.loads(date_temp.get(str(i))))
                    print(i-1101+keep+1,"編號:",temp[0],"名稱:",temp[1],"TEL :",temp[2])
                    
                if number==1 or number==2:
                    print(" 選擇幾號修改 :",end="")
                    number=(input())
                    while number.isdigit()!=True :
                            print(" 選擇幾號修改 :",end='')
                            number=(input())
                    number=int(number)
                    if number<=keep:
                        temp=(json.loads(date_temp.get(str(number+1000))))
                        print("修改名稱為:","(",temp[1],")",end="")
                        temp[1]=input()
                        print("修改TEL 為:","(",temp[2],")",end="")
                        temp[2]=input()
                        lable=str(number+1000)
                        temp=json.dumps(temp)
                        #print(path_goods,lable,temp)
                        update(path_goods,lable,temp)
                        print("update")
                    
                    
                    else:
                        temp=(json.loads(date_temp.get(str(number+1100-keep))))
                        print("修改名稱為:","(",temp[1],")",end="")
                        temp[1]=input()
                        print("修改TEL 為:","(",temp[2],")",end="")
                        temp[2]=input()
                        lable=str(number+1100-keep)
                        temp=json.dumps(temp)
                        #print(path_goods,lable,temp)
                        update(path_goods,lable,temp)
                        print("update")
                
                if number==5 or number==6:
                    print(" 選擇刪除幾號:",end="")
                    number=(input())
                    while number.isdigit()!=True :
                            print(" 選擇刪除幾號:",end='')
                            number=(input())
                    number=int(number)
                    if number<=keep:
                        temp=(json.loads(date_temp.get(str(number+1000))))
                        print("編號:",temp[0],"名稱:",temp[1],"TEL :",temp[2])
                        print(" 1 :確定刪除廠商資料\n","2 :不用刪除廠商資料\n")
                        choose_b=(input())
                        while choose_b.isdigit()!=True :
                            print(" 1 :確定刪除廠商資料\n","2 :不用刪除廠商資料\n",end='')
                            choose_b=(input())
                        choose_b=int(choose_b)
                        lable=str(number+1000)
                        #print(type(lable),type(factory_f))
                        
                        if choose_b == 1:
                            if  int(lable)==int(factory_f):
                                deldate(path_goods+lable)
                                update(path_goods,'f',factory_f-1)
                                print("update")
                            else:      
                                temp=readdate(path_goods+str(factory_f))
                                #print(temp)
                            
                                update(path_goods,lable,temp)
                                deldate(path_goods+str(factory_f))
                                update(path_goods,'f',factory_f-1)
                                print("update")
                    else:
                        
                        temp=json.loads(date_temp.get(str((number+1100-keep))))
                        print("編號:",temp[0],"名稱:",temp[1],"TEL :",temp[2])
                        print(" 1 :確定刪除客戶資料\n","2 :不用刪除客戶資料\n")
                        choose_b=(input())
                        while choose_b.isdigit()!=True :
                            print(" 1 :確定刪除廠商資料\n","2 :不用刪除廠商資料\n",end='')
                            choose_b=(input())
                        choose_=int(choose_b)
                        lable=str(number+1100-keep)
                        #print((lable),(business_b))
                        
                        if choose_b == 1:
                            if  int(lable)==int(business_b):
                                deldate(path_goods+lable)
                                update(path_goods,'b',business_b-1)
                                print("update")
                            else:      
                                temp=readdate(path_goods+str(business_b))
                                #print(temp)
                            
                                update(path_goods,lable,temp)
                                deldate(path_goods+str(business_b))
                                update(path_goods,'b',business_b-1)
                                print("update")

            
            if number==3:
                temp=[]
                temp.append("fa"+str(int(factory_f)+1))
                print("請輸入廠商名稱   : ",end=" ")
                
                temp.append(input())
                print("請輸入廠商電話   : ",end=" ")
                
                temp.append(input())
                print("編號","fa"+str(int(factory_f)+1),"   廠商名稱",temp[1],"廠商電話",temp[2])
                #print(temp)
                temp=json.dumps(temp)
                print(" 1 :確定增加廠商資料\n","2 :不用增加廠商資料\n")
                choose_b=(input())
                while choose_b.isdigit()!=True :
                            print(" 1 :確定刪除廠商資料\n","2 :不用刪除廠商資料\n",end='')
                            choose_b=(input())
                choose_b=int(choose_b)
                if choose_b == 1:
                    update(path_goods,str(int(factory_f)+1),temp)
                    update(path_goods,'f',int(factory_f)+1)
                    print("update")
            
            if number==4:
                temp=[]
                temp.append("bu"+str(int(business_b)+1))
                print("請輸入客戶名稱   : ",end=" ")
                
                temp.append(input())
                print("請輸入客戶電話   : ",end=" ")
                
                temp.append(input())
                print("編號","bu"+str(int(business_b)+1),"   客戶名稱",temp[1],"客戶電話",temp[2])
                #print(temp)
                temp=json.dumps(temp)
                print(" 1 :確定增加客戶資料\n","2 :不用增加客戶資料\n")
                choose_b=(input())
                while choose_b.isdigit()!=True :
                            print(" 1 :確定增加客戶資料\n","2 :不用增加客戶資料\n",end='')
                            choose_b=(input())
                choose_b=int(choose_b)
                if choose_b == 1:
                    update(path_goods,str(int(business_b)+1),temp)
                    update(path_goods,'b',int(business_b)+1)
                    print("update")       
            else:
                number=88 